import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEV_IPS = ['info', 'low', 'medium', 'high', 'critical']
RISK_APP = ['information', 'low', 'medium', 'high', 'elevated']


def procesar_logs_legitimo(ruta_legitimo):
    """Lee todos los logs de tráfico legítimo y los aglutina en una única macro-fila."""
    sessions_any = set()
    sessions_ips = set()
    sessions_app = set()

    total_alerts = 0
    unique_attackids = set()
    unique_appids = set()

    ips_by_sev = {s: {"alerts": 0, "ids": set(), "sessions": set(), "session_ids": {}} for s in SEV_IPS}
    app_by_risk = {r: {"alerts": 0, "ids": set(), "sessions": set(), "session_ids": {}} for r in RISK_APP}

    global_session_events = {}
    archivos_procesados = 0

    for root, dirs, files in os.walk(ruta_legitimo):
        for f_name in files:
            if f_name.endswith('.log') or f_name.endswith('.txt'):
                archivos_procesados += 1
                ruta_archivo = os.path.join(root, f_name)

                # Reiniciamos el devid por cada fichero para evitar arrastres
                devid_actual = None

                with open(ruta_archivo, 'r', errors='ignore') as f:
                    for line in f:
                        if re.search(r'type=["\']?traffic["\']?', line, re.IGNORECASE):
                            continue

                        match_session = re.search(r'sessionid=[\'"]?(\d+)', line, re.IGNORECASE)
                        match_attackid = re.search(r'attackid=[\'"]?(\d+)', line, re.IGNORECASE)
                        match_sev = re.search(r'severity=[\'"]?([a-zA-Z]+)', line, re.IGNORECASE)
                        match_appid = re.search(r'appid=[\'"]?(\d+)', line, re.IGNORECASE)
                        match_risk = re.search(r'apprisk=[\'"]?([a-zA-Z]+)', line, re.IGNORECASE)

                        # FIX: Clave de sesión GLOBAL por dispositivo
                        match_devid = re.search(r'devid=[\'"]?([^"\'\s,]+)', line, re.IGNORECASE)
                        if match_devid:
                            devid_actual = match_devid.group(1)

                        # Si hay sesión, la prefijamos con el devid (si existe) o con el nombre del fichero
                        session_id = None
                        if match_session:
                            if devid_actual:
                                session_id = f"{devid_actual}_{match_session.group(1)}"
                            else:
                                session_id = f"{f_name}_{match_session.group(1)}"

                        atk_id = match_attackid.group(1) if match_attackid else None
                        app_id = match_appid.group(1) if match_appid else None

                        # --- CONTEO (Sin importar si la acción es PASS o DROP) ---
                        if atk_id or app_id:
                            total_alerts += 1
                            if session_id:
                                sessions_any.add(session_id)
                                if session_id not in global_session_events:
                                    global_session_events[session_id] = set()
                                global_session_events[session_id].add((atk_id, app_id))

                        # --- PROCESADO IPS ---
                        if match_attackid and match_sev:
                            sev = match_sev.group(1).lower()
                            if sev in SEV_IPS:
                                unique_attackids.add(atk_id)
                                if session_id:
                                    sessions_ips.add(session_id)

                                ips_by_sev[sev]["alerts"] += 1
                                ips_by_sev[sev]["ids"].add(atk_id)

                                if session_id:
                                    ips_by_sev[sev]["sessions"].add(session_id)
                                    if session_id not in ips_by_sev[sev]["session_ids"]:
                                        ips_by_sev[sev]["session_ids"][session_id] = set()
                                    ips_by_sev[sev]["session_ids"][session_id].add(atk_id)

                        # --- PROCESADO APP ---
                        if match_appid and match_risk:
                            risk = match_risk.group(1).lower()
                            if risk in RISK_APP:
                                unique_appids.add(app_id)
                                if session_id:
                                    sessions_app.add(session_id)

                                app_by_risk[risk]["alerts"] += 1
                                app_by_risk[risk]["ids"].add(app_id)

                                if session_id:
                                    app_by_risk[risk]["sessions"].add(session_id)
                                    if session_id not in app_by_risk[risk]["session_ids"]:
                                        app_by_risk[risk]["session_ids"][session_id] = set()
                                    app_by_risk[risk]["session_ids"][session_id].add(app_id)

    print(f"  -> Se han procesado {archivos_procesados} archivos de tráfico legítimo.")

    # Cálculo métrica consolidada global
    global_diff_alerts = sum(len(s) for s in global_session_events.values())

    # Construcción de la fila de resultados (Con columna 1 añadida para alinear todo)
    res = [
        "Legítimo_Agregado",
        total_alerts,
        global_diff_alerts,
        len(unique_attackids),
        len(unique_appids),
        len(sessions_any),
        len(sessions_ips),
        len(sessions_app),
        len(sessions_ips.union(sessions_app)),
        ""  # Comentarios
    ]

    for sev in SEV_IPS:
        datos = ips_by_sev[sev]
        if datos["alerts"] > 0:
            res.extend([
                datos["alerts"],
                datos["alerts"],
                sum(len(s) for s in datos["session_ids"].values()),
                ", ".join(sorted(list(datos["ids"]))),
                len(datos["ids"]),
                len(datos["sessions"])
            ])
        else:
            res.extend([0, 0, 0, 0, 0, 0])

    for risk in RISK_APP:
        datos = app_by_risk[risk]
        if datos["alerts"] > 0:
            res.extend([
                datos["alerts"],
                datos["alerts"],
                sum(len(s) for s in datos["session_ids"].values()),
                ", ".join(sorted(list(datos["ids"]))),
                len(datos["ids"]),
                len(datos["sessions"])
            ])
        else:
            res.extend([0, 0, 0, 0, 0, 0])

    return res


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)

    print("=== PROCESADOR DE TRÁFICO LEGÍTIMO: FORTIGATE (Metodología TFG + Fix Alineación) ===")

    ruta_legitimo = os.path.join("FG", "Legitimo")
    if not os.path.exists(ruta_legitimo):
        print(f"[!] ERROR: No existe la ruta {ruta_legitimo}.")
        return

    plantilla = "plantillaFG_legitimo.xlsx"
    nombre_salida = "Resultados_Legitimo_FG.xlsx"

    if not os.path.exists(plantilla):
        print(f"[!] ERROR: No se encuentra '{plantilla}'.")
        return

    print(f"[*] Escaneando logs en {ruta_legitimo} y generando {nombre_salida}...")

    valores_fila = procesar_logs_legitimo(ruta_legitimo)

    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active

    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    fill_ips_color = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill_app_color = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')

    row_num = 3
    for col_num, val in enumerate(valores_fila, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.border = border_thin
        cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')

        # Aplicamos colores correctamente desde la columna 11 (IPS) y 41 (APP)
        if col_num >= 11 and col_num <= 40:
            cell.fill = fill_ips_color
        elif col_num >= 41:
            cell.fill = fill_app_color

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 35
    for c in range(1, 71):
        ws.column_dimensions[get_column_letter(c)].width = 16

    wb.save(nombre_salida)
    print(f"✅ ¡Hecho! Archivo guardado ordenadamente en: {nombre_salida}")


if __name__ == '__main__':
    main()