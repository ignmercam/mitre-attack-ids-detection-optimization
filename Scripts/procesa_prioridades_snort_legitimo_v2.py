import os
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from collections import Counter


def main():
    # --- FIJAR DIRECTORIO DE TRABAJO AL DEL SCRIPT ---
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)

    print("=== TFG: PROCESAMIENTO DE PRIORIDADES (TRÁFICO LEGÍTIMO / FP) ===\n")
    print(f"[*] Directorio de trabajo: {base_dir}")

    ruta_plantilla = "plantilla_prioridades_legitimo.xlsx"
    ruta_salida = "Resultados_Prioridades_Legitimo.xlsx"

    if not os.path.exists(ruta_plantilla):
        print(f"[!] Error: No se encuentra {ruta_plantilla}")
        return

    # --- 1. MAPEO INTELIGENTE DE COLUMNAS ---
    print("[1/3] Mapeando columnas de la plantilla...")
    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.worksheets[0]

    bloques = {}
    bloque_actual = None

    for col in range(1, ws.max_column + 1):
        val1 = str(ws.cell(row=1, column=col).value or "").strip()

        if val1:
            # Detectar el RSX
            m_rs = re.search(r'RS(\d+)', val1)
            if m_rs:
                rs_num = int(m_rs.group(1))

                # Detectar la prioridad (puede ser un número o 'Any')
                m_pri = re.search(r'Priority\s*=\s*(\d+|Any)', val1, re.IGNORECASE)
                if m_pri:
                    pri_val = m_pri.group(1)
                    current_priority = int(pri_val) if pri_val.isdigit() else 'Any'
                else:
                    current_priority = 'Any'

                bloque_actual = (rs_num, current_priority)
                if bloque_actual not in bloques:
                    bloques[bloque_actual] = {}

        # Leer Fila 2 para ubicar las subcolumnas exactas
        if bloque_actual:
            val2 = str(ws.cell(row=2, column=col).value or "").strip().lower()
            key = None

            # [!] EL ORDEN DE COMPROBACIÓN IMPORTA
            if 'diferentes por flujo' in val2:
                key = 'unique_alerts_per_flow'
            elif 'alertas/sid' in val2:
                key = 'alerts_per_sid'
            elif '#alertas' in val2 or '#alerts' in val2:
                key = 'alerts'
            elif 'flujos' in val2:
                key = 'flows'
            elif 'sin repetición' in val2 or 'sin repeticion' in val2:
                key = 'sids_list'
            elif '#sid' in val2 or '#sids' in val2:
                key = 'sid_count'

            if key:
                bloques[bloque_actual][key] = col

    # --- 2. LECTURA Y CONSOLIDACIÓN DE LOGS DE TRÁFICO LEGÍTIMO ---
    print("[2/3] Escaneando logs de tráfico legítimo (Consolidación Global)...")

    # Estructura: db[prioridad_o_any][rs_num] = {stats}
    db = {'Any': {}}
    archivos_leidos = 0
    carpetas_base = os.listdir('.')

    regex_sid = re.compile(r'\[\*\*\]\s+\[\d+:(\d+):\d+\]')
    regex_prio = re.compile(r'\[Priority:\s*(\d+)\]', re.IGNORECASE)
    regex_flujo = re.compile(r'\{(.*?)\}\s+(\S+)\s+->\s+(\S+)')

    for d in carpetas_base:
        if os.path.isdir(d) and d.startswith('RS'):
            rs_match = re.search(r'RS(\d+)', d)
            if rs_match:
                rs_num = int(rs_match.group(1))
                leg_dir = os.path.join(d, "Legitimo")

                if os.path.exists(leg_dir):
                    if rs_num not in db['Any']:
                        db['Any'][rs_num] = {'alerts': 0, 'sids': Counter(), 'flows': set(), 'flows_sids': {}}

                    for root, _, files in os.walk(leg_dir):
                        for f in files:
                            if f.endswith('.log') or f.endswith('.txt'):
                                archivos_leidos += 1
                                with open(os.path.join(root, f), 'r', errors='ignore') as file:
                                    for line in file:
                                        m_sid = regex_sid.search(line)
                                        m_prio = regex_prio.search(line)
                                        m_flujo = regex_flujo.search(line)

                                        if m_sid and m_prio:
                                            sid = m_sid.group(1)
                                            pri = int(m_prio.group(1))

                                            # --- IDENTIFICACIÓN ÚNICA IDENTICA A OPTIMO_SNORT_GM_V3.PY ---
                                            pcap_leg = f.replace('.log', '').replace('.txt', '')

                                            if m_flujo:
                                                proto = m_flujo.group(1).strip()
                                                src = m_flujo.group(2).strip()
                                                dst = m_flujo.group(3).strip()
                                                ep1, ep2 = sorted([src, dst])
                                                # Formato exacto del optimizador: pcap__proto ep1 <-> ep2
                                                flujo = f"{pcap_leg}__{proto} {ep1} <-> {ep2}"
                                            else:
                                                # Formato exacto del optimizador para líneas sin flujo estándar
                                                flujo = f"{pcap_leg}__{line.strip()}"

                                            # 1. Sumar al bloque 'Any' (Consolidación total)
                                            db['Any'][rs_num]['alerts'] += 1
                                            db['Any'][rs_num]['sids'][sid] += 1

                                            db['Any'][rs_num]['flows'].add(flujo)
                                            if flujo not in db['Any'][rs_num]['flows_sids']:
                                                db['Any'][rs_num]['flows_sids'][flujo] = set()
                                            db['Any'][rs_num]['flows_sids'][flujo].add(sid)

                                            # 2. Sumar al bloque de su prioridad específica
                                            if pri not in db:
                                                db[pri] = {}
                                            if rs_num not in db[pri]:
                                                db[pri][rs_num] = {'alerts': 0, 'sids': Counter(), 'flows': set(),
                                                                   'flows_sids': {}}

                                            db[pri][rs_num]['alerts'] += 1
                                            db[pri][rs_num]['sids'][sid] += 1

                                            db[pri][rs_num]['flows'].add(flujo)
                                            if flujo not in db[pri][rs_num]['flows_sids']:
                                                db[pri][rs_num]['flows_sids'][flujo] = set()
                                            db[pri][rs_num]['flows_sids'][flujo].add(sid)

    print(f"  -> {archivos_leidos} archivos .log de tráfico legítimo procesados.")

    # --- 3. RELLENAR LA PLANTILLA CON CEROS POR DEFECTO ---
    print("[3/3] Volcando métricas de FP en el Excel...")

    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

    row_idx = 3

    # Recorrer todos los bloques mapeados desde la plantilla
    for (rs_num, prio), cols in bloques.items():
        stats = None
        # Comprobar si realmente hubo alertas para esa combinación
        if prio in db and rs_num in db[prio]:
            stats = db[prio][rs_num]

        # Rellenar cada subcolumna garantizando el 0 si está vacía
        if 'alerts' in cols:
            val = stats['alerts'] if stats and stats['alerts'] > 0 else 0
            c = ws.cell(row=row_idx, column=cols['alerts'], value=val)
            c.border, c.alignment = border_thin, alignment

        # NUEVA COLUMNA: Alertas diferentes por flujo
        if 'unique_alerts_per_flow' in cols:
            val = sum(len(sids) for sids in stats['flows_sids'].values()) if stats and stats['alerts'] > 0 else 0
            c = ws.cell(row=row_idx, column=cols['unique_alerts_per_flow'], value=val)
            c.border, c.alignment = border_thin, alignment

        if 'sid_count' in cols:
            val = len(stats['sids']) if stats and stats['alerts'] > 0 else 0
            c = ws.cell(row=row_idx, column=cols['sid_count'], value=val)
            c.border, c.alignment = border_thin, alignment

        if 'flows' in cols:
            val = len(stats['flows']) if stats and stats['alerts'] > 0 else 0
            c = ws.cell(row=row_idx, column=cols['flows'], value=val)
            c.border, c.alignment = border_thin, alignment

        if 'alerts_per_sid' in cols:
            if stats and stats['alerts'] > 0:
                sids_ord = sorted(stats['sids'].items(), key=lambda x: x[1], reverse=True)
                val = ", ".join([f"{s} ({c})" for s, c in sids_ord])
            else:
                val = 0
            c = ws.cell(row=row_idx, column=cols['alerts_per_sid'], value=val)
            c.border, c.alignment = border_thin, alignment

        if 'sids_list' in cols:
            if stats and stats['alerts'] > 0:
                sids_ord = sorted(stats['sids'].items(), key=lambda x: x[1], reverse=True)
                val = ", ".join([str(s) for s, _ in sids_ord])
            else:
                val = 0
            c = ws.cell(row=row_idx, column=cols['sids_list'], value=val)
            c.border, c.alignment = border_thin, alignment

    wb.save(ruta_salida)
    print(f"\n[+] BARRIDO COMPLETADO CON ÉXITO. Resultados guardados en: {ruta_salida}")


if __name__ == "__main__":
    main()