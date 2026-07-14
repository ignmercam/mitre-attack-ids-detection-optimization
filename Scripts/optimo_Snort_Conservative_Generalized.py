import os
import re
import pandas as pd
import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter

def calc_metrics(test_sids, atk_data_pcap, leg_data_flow, total_atk_pcaps, total_leg_flows,
                 atk_alert_counts, leg_alert_counts, atk_diff_counts, leg_diff_counts):
    """Calcula las métricas exactas basadas en los conjuntos de SIDs activos."""
    tp = sum(1 for pcap_sids in atk_data_pcap.values() if not pcap_sids.isdisjoint(test_sids))
    fn = total_atk_pcaps - tp
    tpr = tp / total_atk_pcaps if total_atk_pcaps > 0 else 0.0

    fp = sum(1 for flow_sids in leg_data_flow.values() if not flow_sids.isdisjoint(test_sids))
    tn = total_leg_flows - fp
    tnr = tn / total_leg_flows if total_leg_flows > 0 else 0.0

    gm = math.sqrt(tpr * tnr)

    alerts_tp = sum(atk_alert_counts[sid] for sid in test_sids)
    alerts_fp = sum(leg_alert_counts[sid] for sid in test_sids)

    alerts_tp_diff = sum(atk_diff_counts[sid] for sid in test_sids)
    alerts_fp_diff = sum(leg_diff_counts[sid] for sid in test_sids)

    return tpr, gm, alerts_tp, alerts_tp_diff, alerts_fp, alerts_fp_diff

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)
    print("=== TFG: CONSERVATIVE GENERALIZED PRUNING - LOG PARSER & EXCEL GENERATOR ===\n")

    TOTAL_ATK_PCAPS = 110
    TOTAL_LEG_FLOWS = 78694206

    regex_sid = re.compile(r'\[\*\*\]\s+\[\d+:(\d+):\d+\]')
    regex_prio = re.compile(r'\[Priority:\s*(\d+)\]', re.IGNORECASE)
    regex_flujo = re.compile(r'\{(.*?)\}\s+(\S+)\s+->\s+(\S+)')

    # Detectar carpetas RS de manera dinámica
    carpetas_rs = sorted([d for d in os.listdir('.') if os.path.isdir(d) and d.startswith('RS')],
                         key=lambda x: int(re.findall(r'\d+', x)[0]) if re.findall(r'\d+', x) else 0)

    if not carpetas_rs:
        # Si no detecta carpetas, forzamos la lista para crear la estructura vacía/rellenable
        carpetas_rs = [f"RS{i}" for i in range(1, 10)]

    # Estructura de almacenamiento de resultados indexada por prioridad y por RS
    escenarios = [4, 3, 2, 1]
    resultados_globales = {pri: {} for pri in escenarios}

    for rs in carpetas_rs:
        print(f"[*] Procesando logs de {rs}...")
        atk_master_dict = defaultdict(list)
        atk_master_flows = defaultdict(list)
        leg_master_dict = defaultdict(list)

        # 1. LEER LOGS DE ATAQUES
        ruta_ataques = os.path.join(rs, "Ataques")
        if os.path.exists(ruta_ataques):
            for f in os.listdir(ruta_ataques):
                if f.endswith('.log') or f.endswith('.txt'):
                    pcap_name = f.replace('.log', '').replace('.txt', '')
                    with open(os.path.join(ruta_ataques, f), 'r', errors='ignore') as file:
                        for line in file:
                            m_sid = regex_sid.search(line)
                            m_pri = regex_prio.search(line)
                            if m_sid and m_pri:
                                sid, pri = m_sid.group(1), int(m_pri.group(1))
                                atk_master_dict[pcap_name].append((sid, pri))
                                m_flujo = regex_flujo.search(line)
                                if m_flujo:
                                    proto, src, dst = m_flujo.group(1).strip(), m_flujo.group(2).strip(), m_flujo.group(3).strip()
                                    ep1, ep2 = sorted([src, dst])
                                    flow_id = f"{proto} {ep1} <-> {ep2}"
                                else:
                                    flow_id = line.strip()
                                atk_master_flows[f"{pcap_name}__{flow_id}"].append((sid, pri))

        # 2. LEER LOGS DE TRAFICO LEGITIMO
        ruta_legitimo = os.path.join(rs, "Legitimo")
        if os.path.exists(ruta_legitimo):
            for f in os.listdir(ruta_legitimo):
                if f.endswith('.log') or f.endswith('.txt'):
                    pcap_leg = f.replace('.log', '').replace('.txt', '')
                    with open(os.path.join(ruta_legitimo, f), 'r', errors='ignore') as file:
                        for line in file:
                            m_sid = regex_sid.search(line)
                            m_pri = regex_prio.search(line)
                            if m_sid and m_pri:
                                sid, pri = m_sid.group(1), int(m_pri.group(1))
                                m_flujo = regex_flujo.search(line)
                                if m_flujo:
                                    proto, src, dst = m_flujo.group(1).strip(), m_flujo.group(2).strip(), m_flujo.group(3).strip()
                                    ep1, ep2 = sorted([src, dst])
                                    flow_id = f"{proto} {ep1} <-> {ep2}"
                                else:
                                    flow_id = line.strip()
                                leg_master_dict[f"{pcap_leg}__{flow_id}"].append((sid, pri))

        # 3. EVALUAR POR ESCENARIO DE PRIORIDAD
        for pri_limit in escenarios:
            atk_filtered = defaultdict(set)
            atk_alert_counts = Counter()
            for pcap, eventos in atk_master_dict.items():
                for sid, prio in eventos:
                    if prio <= pri_limit:
                        atk_filtered[pcap].add(sid)
                        atk_alert_counts[sid] += 1

            atk_filtered_flows = defaultdict(set)
            for flow_id, eventos in atk_master_flows.items():
                for sid, prio in eventos:
                    if prio <= pri_limit:
                        atk_filtered_flows[flow_id].add(sid)

            leg_filtered = defaultdict(set)
            leg_alert_counts = Counter()
            for flow_id, eventos in leg_master_dict.items():
                for sid, prio in eventos:
                    if prio <= pri_limit:
                        leg_filtered[flow_id].add(sid)
                        leg_alert_counts[sid] += 1

            atk_diff_counts = Counter()
            for sids in atk_filtered_flows.values():
                for sid in sids: atk_diff_counts[sid] += 1

            leg_diff_counts = Counter()
            for sids in leg_filtered.values():
                for sid in sids: leg_diff_counts[sid] += 1

            sids_ataque = set(sid for sids in atk_filtered.values() for sid in sids)
            sids_legitimo = set(sid for sids in leg_filtered.values() for sid in sids)
            sids_activos = sids_ataque.union(sids_legitimo)

            # Lógica de Conservative Pruning Generalizado: Cero Tolerancia FP
            sids_cons_gen = sids_activos - sids_legitimo
            removed_sids = sids_legitimo

            tpr, gm, a_tp, a_tp_d, a_fp, a_fp_d = calc_metrics(
                sids_cons_gen, atk_filtered, leg_filtered, TOTAL_ATK_PCAPS, TOTAL_LEG_FLOWS,
                atk_alert_counts, leg_alert_counts, atk_diff_counts, leg_diff_counts
            )

            sorted_removed = sorted(list(removed_sids), key=lambda x: int(x) if str(x).isdigit() else x)
            removed_str = ", ".join(str(s) for s in sorted_removed) if sorted_removed else "None"
            removed_count = len(sorted_removed)
            a_tot = a_tp + a_fp
            usab = round(a_fp / a_tp, 4) if a_tp > 0 else 0.0

            resultados_globales[pri_limit][rs] = {
                'gm': gm,
                'removed_sids': removed_str,
                'removed_count': removed_count,
                'a_tot': a_tot,
                'tpr': f"{tpr * 100:.2f}%",
                'a_tp': a_tp,
                'a_tp_diff': a_tp_d,
                'a_fp': a_fp,
                'a_fp_diff': a_fp_d,
                'usab': usab
            }

    # =========================================================================
    # GENERACIÓN DEL LIBRO DE EXCEL DESDE CERO (SIN PLANTILLAS EXTERNAS)
    # =========================================================================
    print("\n[*] Generando archivo Excel con diseño profesional...")
    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Definición de Estilos
    font_title = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_metric_header = Font(name='Segoe UI', size=11, bold=True, color='1F497D')
    font_bold = Font(name='Segoe UI', size=10, bold=True)
    font_regular = Font(name='Segoe UI', size=10)
    
    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid') # Azul Navy oscuro
    fill_sub_header = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid') # Azul muy claro desgajado
    
    border_thin = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='thin', color='A6A6A6')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    metrics_layout = [
        ("GM", 'gm'),
        ("SIDs/AttackID/ThreatID o AppID eliminados", 'removed_sids'),
        ("#SIDs/AttackID/ThreatID o AppID eliminados", 'removed_count'),
        ("#Alertas total", 'a_tot'),
        ("%Detección_pcaps", 'tpr'),
        ("#Alertas_TP", 'a_tp'),
        ("#Alertas TP (Different/Flow)", 'a_tp_diff'),
        ("#Alertas_FP", 'a_fp'),
        ("#Alertas FP (Different/Flow)", 'a_fp_diff'),
        ("Usabilidad", 'usab')
    ]

    for pri in escenarios:
        ws = wb.create_sheet(title=f"Pri <= {pri}")
        ws.sheet_view.showGridLines = True

        # Fila 1: Cabecera Principal
        ws.cell(row=1, column=1, value="Metric").font = font_title
        ws.cell(row=1, column=1).fill = fill_header
        ws.cell(row=1, column=1).alignment = align_center
        ws.cell(row=1, column=1).border = border_thin

        for col_idx, rs_name in enumerate(carpetas_rs, start=2):
            cell = ws.cell(row=1, column=col_idx, value=rs_name)
            cell.font = font_title
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin

        # Fila 2: Título de la sección
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(carpetas_rs) + 1)
        title_cell = ws.cell(row=2, column=1, value="Conservative Generalized Pruning")
        title_cell.font = font_metric_header
        title_cell.fill = fill_sub_header
        title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        title_cell.border = border_thin
        
        # Aplicar borde a las celdas combinadas de la fila 2
        for col_idx in range(1, len(carpetas_rs) + 2):
            ws.cell(row=2, column=col_idx).border = border_thin

        # Rellenar filas de métricas
        for row_offset, (label, key) in enumerate(metrics_layout, start=3):
            lbl_cell = ws.cell(row=row_offset, column=1, value=label)
            lbl_cell.font = font_bold
            lbl_cell.border = border_thin
            lbl_cell.alignment = align_left

            for col_idx, rs_name in enumerate(carpetas_rs, start=2):
                val = ""
                if rs_name in resultados_globales[pri]:
                    val = resultados_globales[pri][rs_name].get(key, "")
                
                val_cell = ws.cell(row=row_offset, column=col_idx, value=val)
                val_cell.font = font_regular
                val_cell.border = border_thin
                
                # Alineación y formato condicionado por tipo de datos
                if key == 'removed_sids':
                    val_cell.alignment = align_left
                else:
                    val_cell.alignment = align_center

        # Autoajustar anchos de columna de forma limpia
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Si está combinada la fila 2, ignoramos su longitud para no distorsionar el ancho
                if cell.row == 2: continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Margen de seguridad razonable
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
        # Forzar columna A más ancha para las etiquetas de métricas
        ws.column_dimensions['A'].width = 42

    output_filename = "Snort_Conservative_Generalized_Pruning.xlsx"
    wb.save(output_filename)
    print(f"\n✅ EXCEL PROCESADO Y GUARDADO CON ÉXITO: {output_filename}")

if __name__ == "__main__":
    main()
