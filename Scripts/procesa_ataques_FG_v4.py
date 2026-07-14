import pandas as pd
import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Listas de severidades en el orden exacto solicitado para las columnas
SEV_IPS = ['info', 'low', 'medium', 'high', 'critical']
RISK_APP = ['information', 'low', 'medium', 'high', 'elevated']


def analizar_fichero_log(ruta_archivo):
    """Parsea un archivo .log de FortiGate y extrae métricas detalladas por flujos."""
    sessions_any = set()
    sessions_ips = set()
    sessions_app = set()

    total_alerts = 0
    unique_attackids = set()
    unique_appids = set()

    ips_by_sev = {s: {"alerts": 0, "ids": {}, "sessions": set(), "session_ids": {}} for s in SEV_IPS}
    app_by_risk = {r: {"alerts": 0, "ids": {}, "sessions": set(), "session_ids": {}} for r in RISK_APP}

    # Tracking global para la métrica consolidada (AttackID_o_AppID)
    global_session_events = {}

    with open(ruta_archivo, 'r', errors='ignore') as f:
        for line in f:
            if re.search(r'type=["\']?traffic["\']?', line, re.IGNORECASE):
                continue
            match_session = re.search(r'sessionid=[\'"]?(\d+)', line, re.IGNORECASE)
            session_id = match_session.group(1) if match_session else None

            if session_id and session_id not in global_session_events:
                global_session_events[session_id] = set()

            match_attackid = re.search(r'attackid=[\'"]?(\d+)', line, re.IGNORECASE)
            match_sev = re.search(r'severity=[\'"]?([a-zA-Z]+)', line, re.IGNORECASE)
            match_appid = re.search(r'appid=[\'"]?(\d+)', line, re.IGNORECASE)
            match_risk = re.search(r'apprisk=[\'"]?([a-zA-Z]+)', line, re.IGNORECASE)

            is_ips = False
            is_app = False

            # --- PROCESADO IPS ---
            if match_attackid and match_sev:
                attackid = match_attackid.group(1)
                sev = match_sev.group(1).lower()

                if sev in SEV_IPS:
                    is_ips = True
                    sessions_ips.add(session_id)
                    unique_attackids.add(attackid)

                    if session_id:
                        global_session_events[session_id].add(('ips', attackid))

                    ips_by_sev[sev]["alerts"] += 1
                    ips_by_sev[sev]["sessions"].add(session_id)
                    ips_by_sev[sev]["ids"][attackid] = ips_by_sev[sev]["ids"].get(attackid, 0) + 1

                    # Tracking específico por sesión para métrica Diferentes/Flujo en IPS
                    if session_id:
                        if session_id not in ips_by_sev[sev]["session_ids"]:
                            ips_by_sev[sev]["session_ids"][session_id] = set()
                        ips_by_sev[sev]["session_ids"][session_id].add(attackid)

            # --- PROCESADO APP ---
            if match_appid and match_risk:
                appid = match_appid.group(1)
                risk = match_risk.group(1).lower()

                if risk in RISK_APP:
                    is_app = True
                    sessions_app.add(session_id)
                    unique_appids.add(appid)

                    if session_id:
                        global_session_events[session_id].add(('app', appid))

                    app_by_risk[risk]["alerts"] += 1
                    app_by_risk[risk]["sessions"].add(session_id)
                    app_by_risk[risk]["ids"][appid] = app_by_risk[risk]["ids"].get(appid, 0) + 1

                    # Tracking específico por sesión para métrica Diferentes/Flujo en APP
                    if session_id:
                        if session_id not in app_by_risk[risk]["session_ids"]:
                            app_by_risk[risk]["session_ids"][session_id] = set()
                        app_by_risk[risk]["session_ids"][session_id].add(appid)

            if is_ips or is_app:
                total_alerts += 1
                if session_id:
                    sessions_any.add(session_id)

    archivo_limpio = os.path.basename(ruta_archivo).replace(".log", "").replace(".txt", "")

    # Cálculo métrica consolidada global (AttackID_o_AppID diferentes por flujo)
    global_diff_alerts = sum(len(s) for s in global_session_events.values())

    res = [
        archivo_limpio,
        total_alerts,
        global_diff_alerts,  # <-- NUEVA COLUMNA GLOBAL
        len(unique_attackids),
        len(unique_appids),
        len(sessions_any),
        len(sessions_ips),
        len(sessions_app),
        len(sessions_ips.union(sessions_app)),
        ""  # Comentarios
    ]

    for sev in SEV_IPS:
        datos = ips_by_sev[sev]
        if datos["alerts"] > 0:
            ordenados = sorted(datos["ids"].items(), key=lambda x: x[1], reverse=True)
            alertas_por_id_str = ", ".join([f"{k} ({v})" for k, v in ordenados])
            ids_str = ", ".join([k for k, v in ordenados])

            # Cálculo Diferentes/Flujo interno para este nivel IPS
            diff_alerts_sev = sum(len(s) for s in datos["session_ids"].values())

            res.extend([
                datos["alerts"],
                alertas_por_id_str,
                diff_alerts_sev,  # <-- NUEVA COLUMNA ESPECÍFICA IPS
                ids_str,
                len(datos["ids"]),
                len(datos["sessions"])
            ])
        else:
            res.extend([0, 0, 0, 0, 0, 0])

    for risk in RISK_APP:
        datos = app_by_risk[risk]
        if datos["alerts"] > 0:
            ordenados = sorted(datos["ids"].items(), key=lambda x: x[1], reverse=True)
            alertas_por_id_str = ", ".join([f"{k} ({v})" for k, v in ordenados])
            ids_str = ", ".join([k for k, v in ordenados])

            # Cálculo Diferentes/Flujo interno para este nivel APP
            diff_alerts_risk = sum(len(s) for s in datos["session_ids"].values())

            res.extend([
                datos["alerts"],
                alertas_por_id_str,
                diff_alerts_risk,  # <-- NUEVA COLUMNA ESPECÍFICA APP
                ids_str,
                len(datos["ids"]),
                len(datos["sessions"])
            ])
        else:
            res.extend([0, 0, 0, 0, 0, 0])

    return res


def cargar_orden(ruta_orden):
    try:
        df = pd.read_excel(ruta_orden, header=None)
        if 'ataque' in str(df.iloc[0, 0]).lower():
            return df.iloc[1:, 0].dropna().astype(str).str.strip().tolist()
        return df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
    except Exception as e:
        print(f"[!] Error al cargar orden.xlsx: {e}")
        return []


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)

    print("=== PROCESADOR DE LOGS: FORTIGATE ===")

    # RUTA DIRECTA (Sin RSX)
    ruta_ataques = os.path.join("FG", "Ataques")
    if not os.path.exists(ruta_ataques):
        print(f"[!] ERROR: No existe la ruta {ruta_ataques}.")
        return

    archivos_pcap = [f for f in os.listdir(ruta_ataques) if f.endswith('.log') or f.endswith('.txt')]
    if not archivos_pcap:
        print(f"[-] No hay archivos .log o .txt en {ruta_ataques}")
        return

    orden = cargar_orden("orden.xlsx")
    archivos_ordenados = []

    if orden:
        for pcap_esperado in orden:
            encontrado = None
            for fname in archivos_pcap:
                if fname.startswith(pcap_esperado):
                    encontrado = fname
                    break
            if encontrado:
                archivos_ordenados.append(os.path.join(ruta_ataques, encontrado))
    else:
        archivos_ordenados = [os.path.join(ruta_ataques, f) for f in sorted(archivos_pcap)]

    plantilla = "plantillaFG.xlsx"
    nombre_salida = "Resultados_Ataques_FG.xlsx"

    if not os.path.exists(plantilla):
        print(f"[!] ERROR: No se encuentra '{plantilla}'.")
        return

    print(f"[*] Procesando logs en {ruta_ataques} y generando {nombre_salida}...")
    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active

    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    fill_ips_color = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill_app_color = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')

    row_num = 3
    for archivo in archivos_ordenados:
        valores_fila = analizar_fichero_log(archivo)

        for col_num, val in enumerate(valores_fila, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = border_thin
            cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')

            # Aplicar colores basados en las 70 columnas (10 globales, 30 IPS, 30 APP)
            if col_num >= 11 and col_num <= 40:
                cell.fill = fill_ips_color
            elif col_num >= 41:
                cell.fill = fill_app_color

        row_num += 1

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 35
    for c in range(1, 71):
        ws.column_dimensions[get_column_letter(c)].width = 16

    wb.save(nombre_salida)
    print(f"✅ ¡Hecho! Archivo guardado en: {nombre_salida}")


if __name__ == '__main__':
    main()