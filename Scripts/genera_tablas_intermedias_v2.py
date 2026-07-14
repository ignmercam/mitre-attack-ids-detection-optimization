import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from collections import defaultdict

# --- MAPEOS DE SEVERIDAD ---
FG_SEV_MAP = {'info': 1, 'low': 2, 'medium': 3, 'high': 4, 'critical': 5}
FG_RISK_MAP = {'information': 1, 'low': 2, 'medium': 3, 'high': 4, 'elevated': 5}


def map_fg_ips(sev):
    return FG_SEV_MAP.get(str(sev).strip().lower(), 0)


def map_fg_app(risk):
    return FG_RISK_MAP.get(str(risk).strip().lower(), 0)


def get_base_tech(name):
    """Extrae la técnica TXXXX de cualquier string."""
    if not name: return None
    match = re.search(r'(T\d{4})', str(name))
    return match.group(1) if match else None


def parse_snort_raw(tecnicas_validas):
    """Escanea las carpetas RSX/Ataques y agrupa alertas. Cuenta los logs SOLO por cada RS."""
    snort_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    snort_totals = defaultdict(lambda: defaultdict(int))  # Cuenta logs: [tecnica][rs_num] = total_logs

    regex_sid = re.compile(r'\[\*\*\]\s+\[\d+:(\d+):\d+\]')
    regex_prio = re.compile(r'\[Priority:\s*(\d+)\]', re.IGNORECASE)
    regex_flujo = re.compile(r'\{(.*?)\}\s+(\S+)\s+->\s+(\S+)')

    carpetas = [d for d in os.listdir('.') if os.path.isdir(d) and d.startswith('RS')]

    for d in carpetas:
        m_rs = re.search(r'RS(\d+)', d)
        if not m_rs: continue
        rs_num = int(m_rs.group(1))

        atk_dir = os.path.join(d, "Ataques")
        if not os.path.exists(atk_dir): continue

        for root, _, files in os.walk(atk_dir):
            for f in files:
                if f.endswith('.log') or f.endswith('.txt'):
                    pcap_name = f.replace('.log', '').replace('.txt', '')
                    tech = get_base_tech(pcap_name)

                    if tech in tecnicas_validas:
                        # SUMAMOS 1 AL TOTAL DE PCAPS DE ESTA TÉCNICA *SOLO EN ESTE RS*
                        snort_totals[tech][rs_num] += 1

                        with open(os.path.join(root, f), 'r', errors='ignore') as file:
                            for line in file:
                                m_sid = regex_sid.search(line)
                                m_pri = regex_prio.search(line)
                                m_flujo = regex_flujo.search(line)

                                if m_sid and m_pri:
                                    sid = int(m_sid.group(1))
                                    pri = int(m_pri.group(1))

                                    if m_flujo:
                                        proto = m_flujo.group(1).strip()
                                        ep1, ep2 = sorted([m_flujo.group(2).strip(), m_flujo.group(3).strip()])
                                        flujo_id = f"{proto} {ep1} <-> {ep2}"
                                    else:
                                        flujo_id = f"{line.strip()}"

                                    snort_data[tech][rs_num][(pcap_name, flujo_id)].append((sid, pri))
    return snort_data, snort_totals


def parse_fw_raw(engine_type, tecnicas_validas):
    """Escanea FG o PA, agrupa alertas y cuenta los logs SOLO para este Firewall."""
    fw_data = defaultdict(lambda: defaultdict(list))
    fw_totals = defaultdict(int)  # Cuenta logs: [tecnica] = total_logs

    base_path = os.path.join(engine_type, "Ataques")

    if not os.path.exists(base_path):
        return fw_data, fw_totals

    re_sess = re.compile(r'sessionid=[\'"]?(\d+)', re.IGNORECASE) if engine_type == 'FG' else re.compile(
        r'sessionid="?(\d+)"?', re.IGNORECASE)

    for root, _, files in os.walk(base_path):
        for f in files:
            if f.endswith('.log') or f.endswith('.txt'):
                pcap_name = f.replace('.log', '').replace('.txt', '')
                tech = get_base_tech(pcap_name)

                if tech in tecnicas_validas:
                    # SUMAMOS 1 AL TOTAL DE PCAPS DE ESTA TÉCNICA *SOLO EN ESTE FIREWALL*
                    fw_totals[tech] += 1

                    with open(os.path.join(root, f), 'r', errors='ignore') as file:
                        for line in file:
                            if ('type="traffic"' in line) or ('type=traffic' in line) or (
                                    'TRAFFIC' in line and engine_type == 'PA'):
                                continue

                            match_session = re_sess.search(line)
                            session_id = f"sess_{match_session.group(1)}" if match_session else f"raw_{line.strip()}"

                            ips_val, ips_level = None, 0
                            app_val, app_level = None, 0

                            if engine_type == 'FG':
                                m_atk = re.search(r'attackid=(\d+)', line)
                                m_sev = re.search(r'severity="([^"]+)"', line)
                                if m_atk and m_sev:
                                    ips_val, ips_level = m_atk.group(1), map_fg_ips(m_sev.group(1))

                                m_app = re.search(r'appid=(\d+)', line)
                                m_risk = re.search(r'apprisk="([^"]+)"', line)
                                if m_app and m_risk:
                                    app_val, app_level = m_app.group(1), map_fg_app(m_risk.group(1))

                            elif engine_type == 'PA':
                                m_thr = re.search(r'threat_id="?([^"\s]+)"?', line)
                                m_sev = re.search(r'severity_number="?(\d+)"?', line)
                                if m_thr and m_sev:
                                    ips_val, ips_level = m_thr.group(1), int(m_sev.group(1))
                                    if '(' in ips_val:
                                        ips_val = re.search(r'\((\d+)\)', ips_val).group(1)

                                m_app = re.search(r'(?:appid|app)="?([^"\s]+)"?', line)
                                m_risk = re.search(r'risk_of_app="?(\d+)"?', line)
                                if m_app and m_risk:
                                    app_val, app_level = m_app.group(1), int(m_risk.group(1))

                            if ips_val or app_val:
                                fw_data[tech][(pcap_name, session_id)].append((ips_val, ips_level, app_val, app_level))
    return fw_data, fw_totals


def update_headers_for_severity(ws, min_sev):
    """Actualiza la Fila 1 para mostrar la severidad correcta en FG y PA."""
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value or "")
        if 'severity >=' in val or 'risk_of_app >=' in val or 'apprisk >=' in val or 'severity_number >=' in val:
            new_val = re.sub(r'(>=)\s*\d+', rf'\g<1>{min_sev}', val)
            ws.cell(row=1, column=col, value=new_val)


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)
    print("=== TFG: GENERADOR DE TABLAS BASE DE ATAQUES (POR TÉCNICA) ===\n")

    ruta_plantilla = "plantilla_intermedia.xlsx"
    ruta_tecnicas = "tecnicas.xlsx"

    if not os.path.exists(ruta_plantilla) or not os.path.exists(ruta_tecnicas):
        print(f"[!] Error: Faltan archivos. Asegúrate de tener '{ruta_plantilla}' y '{ruta_tecnicas}'.")
        return

    # --- 1. LEER TECNICAS ---
    print(f"[*] Leyendo lista maestra desde {ruta_tecnicas}...")
    df_tech = pd.read_excel(ruta_tecnicas, header=None)
    tecnicas_lista = df_tech[0].dropna().astype(str).str.strip().tolist()

    tecnicas_filas = {}
    fila_actual = 3
    for t in tecnicas_lista:
        base_tech = get_base_tech(t)
        if base_tech and base_tech not in tecnicas_filas:
            tecnicas_filas[base_tech] = fila_actual
            fila_actual += 1

    # --- 2. MAPEO BLINDADO DE COLUMNAS ---
    wb_out = openpyxl.load_workbook(ruta_plantilla)
    ws_temp = wb_out.worksheets[0]

    bloques = {'RS': {}, 'FG_IPS': {}, 'FG_ALL': {}, 'PA_IPS': {}, 'PA_ALL': {}}
    current_blk = None
    pct_cols = set()

    for col in range(1, ws_temp.max_column + 1):
        val1 = str(ws_temp.cell(row=1, column=col).value or "").strip()

        if 'RS' in val1:
            rs_num = int(re.search(r'RS(\d+)', val1).group(1))
            if rs_num not in bloques['RS']: bloques['RS'][rs_num] = {}
            current_blk = bloques['RS'][rs_num]
        elif 'FG' in val1:
            current_blk = bloques['FG_ALL'] if '||' in val1 else bloques['FG_IPS']
        elif 'PA' in val1:
            current_blk = bloques['PA_ALL'] if '||' in val1 else bloques['PA_IPS']
        elif val1 == "":
            pass
        else:
            current_blk = None

        if current_blk is None: continue

        val2_raw = str(ws_temp.cell(row=2, column=col).value or "")
        val2_norm = re.sub(r'\s+', '', val2_raw.lower())

        if val2_norm == '#alerts':
            current_blk['alerts'] = col
        elif val2_norm == '#flow-correlatedalerts':
            current_blk['alerts_diff'] = col
        elif val2_norm in ['#sid', '#attackid', '#threatid']:
            current_blk['ids'] = col
        elif val2_norm == '#appid':
            current_blk['appid'] = col
        elif val2_norm in ['#attack_flows', '#attack_flowsattackid', '#attack_flowsthreatid']:
            current_blk['attack_flows'] = col
        elif val2_norm in ['%detection_pcaps', '%detection_pcapsattackid', '%detection_pcapsthreatid']:
            current_blk['detection_pcaps'] = col
            pct_cols.add(col)

    # --- 3. RECOLECCIÓN DE LOGS Y CONTEO DINÁMICO DE PCAPS ---
    print("[*] Escaneando Logs RAW de Snort (RS1-RS9)...")
    snort_data, snort_totals = parse_snort_raw(tecnicas_filas.keys())
    print("[*] Escaneando Logs RAW de FortiGate y Palo Alto...")
    fg_data, fg_totals = parse_fw_raw('FG', tecnicas_filas.keys())
    pa_data, pa_totals = parse_fw_raw('PA', tecnicas_filas.keys())
    print("[+] Logs cargados. Procesando escenarios...")

    # --- 4. GENERAR ESCENARIOS ---
    escenarios = [
        ("Caso Any", None, 1),
        ("Snort<=4_FGPA>=1", 4, 1),
        ("Snort<=3_FGPA>=2", 3, 2),
        ("Snort<=2_FGPA>=3", 2, 3),
        ("Snort<=1_FGPA>=4", 1, 4),
        ("FGPA>=5", -1, 5)
    ]

    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    align_center = Alignment(vertical='top', horizontal='center', wrap_text=True)

    for titulo_hoja, max_pri, min_sev in escenarios:
        print(f"  -> Generando hoja: {titulo_hoja}")
        ws = wb_out.copy_worksheet(ws_temp)
        ws.title = titulo_hoja
        update_headers_for_severity(ws, min_sev)

        for tech, row_idx in tecnicas_filas.items():

            # --- A. SNORT ---
            if max_pri != -1:
                for rs_num, cols in bloques['RS'].items():
                    alerts, alerts_diff, attack_flows = 0, 0, 0
                    unique_sids, pcaps_triggered = set(), set()

                    # Dividimos SIEMPRE por el número de logs en ESTA carpeta RS exacta
                    total_pcaps_snort = snort_totals[tech][rs_num]

                    for (pcap_name, flow_id), events in snort_data[tech][rs_num].items():
                        valid_sids = [sid for sid, pri in events if max_pri is None or pri <= max_pri]
                        if valid_sids:
                            attack_flows += 1
                            pcaps_triggered.add(pcap_name)
                            alerts += len(valid_sids)
                            alerts_diff += len(set(valid_sids))
                            unique_sids.update(valid_sids)

                    if alerts > 0:
                        if 'alerts' in cols: ws.cell(row=row_idx, column=cols['alerts'], value=alerts)
                        if 'alerts_diff' in cols: ws.cell(row=row_idx, column=cols['alerts_diff'], value=alerts_diff)
                        if 'ids' in cols: ws.cell(row=row_idx, column=cols['ids'], value=len(unique_sids))
                        if 'attack_flows' in cols: ws.cell(row=row_idx, column=cols['attack_flows'], value=attack_flows)

                    if 'detection_pcaps' in cols:
                        pct = (len(pcaps_triggered) / total_pcaps_snort) if total_pcaps_snort > 0 else 0.0
                        ws.cell(row=row_idx, column=cols['detection_pcaps'], value=pct)
            else:
                for rs_num, cols in bloques['RS'].items():
                    for col_num in cols.values():
                        ws.cell(row=row_idx, column=col_num, value="N/A")

            # --- B. FORTIGATE Y PALO ALTO ---
            for engine in ['FG', 'PA']:
                fw_data_tech = fg_data if engine == 'FG' else pa_data
                fw_totals_tech = fg_totals if engine == 'FG' else pa_totals

                # Dividimos SIEMPRE por el número de logs en ESTE firewall exacto
                total_pcaps_fw = fw_totals_tech[tech]

                alerts_ips, alerts_diff_ips, attack_flows_ips = 0, 0, 0
                unique_ips, pcaps_triggered_ips = set(), set()
                alerts_comp, alerts_diff_comp, attack_flows_comp = 0, 0, 0
                unique_ips_comp, unique_apps_comp, pcaps_triggered_comp = set(), set(), set()

                for (pcap_name, flow_id), events in fw_data_tech[tech].items():
                    valid_ips_events, valid_app_events = [], []

                    for ips_id, ips_sev, app_id, app_sev in events:
                        if ips_id and ips_sev >= min_sev: valid_ips_events.append(ips_id)
                        if app_id and app_sev >= min_sev: valid_app_events.append(app_id)

                    if valid_ips_events:
                        attack_flows_ips += 1
                        pcaps_triggered_ips.add(pcap_name)
                        alerts_ips += len(valid_ips_events)
                        alerts_diff_ips += len(set(valid_ips_events))
                        unique_ips.update(valid_ips_events)

                    if valid_ips_events or valid_app_events:
                        attack_flows_comp += 1
                        pcaps_triggered_comp.add(pcap_name)
                        alerts_comp += len(valid_ips_events) + len(valid_app_events)

                        flow_unique_items = set()
                        flow_unique_items.update(valid_ips_events)
                        flow_unique_items.update(valid_app_events)
                        alerts_diff_comp += len(flow_unique_items)

                        unique_ips_comp.update(valid_ips_events)
                        unique_apps_comp.update(valid_app_events)

                if engine == 'FG':
                    cols_ips = bloques.get('FG_IPS', {})
                    if alerts_ips > 0:
                        if 'alerts' in cols_ips: ws.cell(row=row_idx, column=cols_ips['alerts'], value=alerts_ips)
                        if 'alerts_diff' in cols_ips: ws.cell(row=row_idx, column=cols_ips['alerts_diff'],
                                                              value=alerts_diff_ips)
                        if 'ids' in cols_ips: ws.cell(row=row_idx, column=cols_ips['ids'], value=len(unique_ips))
                        if 'attack_flows' in cols_ips: ws.cell(row=row_idx, column=cols_ips['attack_flows'],
                                                               value=attack_flows_ips)
                    if 'detection_pcaps' in cols_ips:
                        pct_ips = (len(pcaps_triggered_ips) / total_pcaps_fw) if total_pcaps_fw > 0 else 0.0
                        ws.cell(row=row_idx, column=cols_ips['detection_pcaps'], value=pct_ips)

                    cols_all = bloques.get('FG_ALL', {})
                    if alerts_comp > 0:
                        if 'alerts' in cols_all: ws.cell(row=row_idx, column=cols_all['alerts'], value=alerts_comp)
                        if 'alerts_diff' in cols_all: ws.cell(row=row_idx, column=cols_all['alerts_diff'],
                                                              value=alerts_diff_comp)
                        if 'ids' in cols_all: ws.cell(row=row_idx, column=cols_all['ids'], value=len(unique_ips_comp))
                        if 'appid' in cols_all: ws.cell(row=row_idx, column=cols_all['appid'],
                                                        value=len(unique_apps_comp))
                        if 'attack_flows' in cols_all: ws.cell(row=row_idx, column=cols_all['attack_flows'],
                                                               value=attack_flows_comp)
                    if 'detection_pcaps' in cols_all:
                        pct_comp = (len(pcaps_triggered_comp) / total_pcaps_fw) if total_pcaps_fw > 0 else 0.0
                        ws.cell(row=row_idx, column=cols_all['detection_pcaps'], value=pct_comp)

                if engine == 'PA':
                    cols_ips = bloques.get('PA_IPS', {})
                    if alerts_ips > 0:
                        if 'alerts' in cols_ips: ws.cell(row=row_idx, column=cols_ips['alerts'], value=alerts_ips)
                        if 'alerts_diff' in cols_ips: ws.cell(row=row_idx, column=cols_ips['alerts_diff'],
                                                              value=alerts_diff_ips)
                        if 'ids' in cols_ips: ws.cell(row=row_idx, column=cols_ips['ids'], value=len(unique_ips))
                        if 'attack_flows' in cols_ips: ws.cell(row=row_idx, column=cols_ips['attack_flows'],
                                                               value=attack_flows_ips)
                    if 'detection_pcaps' in cols_ips:
                        pct_ips = (len(pcaps_triggered_ips) / total_pcaps_fw) if total_pcaps_fw > 0 else 0.0
                        ws.cell(row=row_idx, column=cols_ips['detection_pcaps'], value=pct_ips)

                    cols_all = bloques.get('PA_ALL', {})
                    if alerts_comp > 0:
                        if 'alerts' in cols_all: ws.cell(row=row_idx, column=cols_all['alerts'], value=alerts_comp)
                        if 'alerts_diff' in cols_all: ws.cell(row=row_idx, column=cols_all['alerts_diff'],
                                                              value=alerts_diff_comp)
                        if 'ids' in cols_all: ws.cell(row=row_idx, column=cols_all['ids'], value=len(unique_ips_comp))
                        if 'appid' in cols_all: ws.cell(row=row_idx, column=cols_all['appid'],
                                                        value=len(unique_apps_comp))
                        if 'attack_flows' in cols_all: ws.cell(row=row_idx, column=cols_all['attack_flows'],
                                                               value=attack_flows_comp)
                    if 'detection_pcaps' in cols_all:
                        pct_comp = (len(pcaps_triggered_comp) / total_pcaps_fw) if total_pcaps_fw > 0 else 0.0
                        ws.cell(row=row_idx, column=cols_all['detection_pcaps'], value=pct_comp)

        # Aplicar el formato visual final a celdas vacías y formatear porcentajes
        for row in range(3, ws.max_row + 1):
            for col in range(2, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                if c.value is None:
                    if max_pri == -1 and any(
                            c.column == col_num for cols in bloques['RS'].values() for col_num in cols.values()):
                        c.value = "N/A"
                    else:
                        c.value = 0

                c.alignment = align_center
                c.border = border_thin

                # ¡Aplicamos formato porcentaje nativo de Excel!
                if col in pct_cols and str(c.value) != "N/A":
                    c.number_format = '0.00%'

    wb_out.remove(ws_temp)

    output_filename = "Resultados_Ataques_Base_Tecnicas.xlsx"
    wb_out.save(output_filename)
    print(f"\n✅ ¡AHORA SÍ! Archivo maestro guardado con métricas y porcentajes perfectos en: {output_filename}")


if __name__ == "__main__":
    main()