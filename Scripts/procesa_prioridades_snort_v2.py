import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from collections import Counter


def cargar_orden(ruta_orden):
    """Carga la lista de ataques desde orden.xlsx para mantener el orden estricto de las filas."""
    try:
        df_orden = pd.read_excel(ruta_orden, header=None)
        # Si la primera celda es una cabecera, la saltamos
        if 'ataque' in str(df_orden.iloc[0, 0]).lower():
            return df_orden.iloc[1:, 0].dropna().astype(str).str.strip().tolist()
        return df_orden.iloc[:, 0].dropna().astype(str).str.strip().tolist()
    except Exception as e:
        print(f"[!] Error al cargar {ruta_orden}: {e}")
        return []


def identificar_pcap(filename, ataques_ordenados_por_longitud):
    """Encuentra a qué ataque pertenece el log, ignorando sufijos."""
    for ataque in ataques_ordenados_por_longitud:
        if filename.startswith(ataque):
            return ataque
    return None


def main():
    # --- FIJAR DIRECTORIO DE TRABAJO AL DEL SCRIPT ---
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)

    print("=== TFG: PROCESAMIENTO DE ALERTAS SNORT POR PRIORIDADES (+ 'ANY' Y ALERTAS/FLUJO) ===")
    print(f"[*] Directorio de trabajo: {base_dir}\n")

    ruta_orden = "orden.xlsx"
    ruta_plantilla = "plantilla_prioridades.xlsx"
    ruta_salida = "Resultados_Prioridades_Completos.xlsx"

    if not os.path.exists(ruta_orden):
        print(f"[!] Archivo no encontrado: {ruta_orden}")
        return
    if not os.path.exists(ruta_plantilla):
        print(f"[!] Archivo de plantilla no encontrado: {ruta_plantilla}")
        return

    orden_list = cargar_orden(ruta_orden)
    # Ordenamos por longitud descendente para que "T1190-Sql_injection" haga match antes que "T1190"
    ataques_para_buscar = sorted(orden_list, key=len, reverse=True)
    print(f"[+] Orden cargado con {len(orden_list)} PCAPs únicos.")

    # --- EXPRESIONES REGULARES ---
    regex_sid = re.compile(r'\[\*\*\]\s+\[\d+:(\d+):\d+\]')
    regex_prio = re.compile(r'\[Priority:\s*(\d+)\]', re.IGNORECASE)
    regex_flujo = re.compile(r'\{(.*?)\}\s+(\S+)\s+->\s+(\S+)')

    # Estructura principal: datos[nombre_ataque][(RS, Prioridad o 'Any')] = {stats}
    datos_completos = {ataque: {} for ataque in orden_list}

    # --- BARRIDO DE CARPETAS RSX ---
    carpetas_rs = sorted([d for d in os.listdir('.') if os.path.isdir(d) and d.startswith('RS')],
                         key=lambda x: int(re.findall(r'\d+', x)[0]) if re.findall(r'\d+', x) else 0)

    for rs in carpetas_rs:
        rs_num = int(re.findall(r'\d+', rs)[0])
        ruta_ataques = os.path.join(rs, "Ataques")

        if not os.path.exists(ruta_ataques):
            continue

        print(f"[*] Escaneando logs en {rs}...")
        for filename in os.listdir(ruta_ataques):
            if not (filename.endswith('.log') or filename.endswith('.txt')):
                continue

            ataque_base = identificar_pcap(filename, ataques_para_buscar)
            if not ataque_base:
                continue

            ruta_log = os.path.join(ruta_ataques, filename)

            try:
                with open(ruta_log, 'r', encoding='utf-8', errors='ignore') as f:
                    for line in f:
                        m_sid = regex_sid.search(line)
                        m_prio = regex_prio.search(line)
                        m_flujo = regex_flujo.search(line)

                        if m_sid and m_prio:
                            sid = m_sid.group(1)
                            prio = int(m_prio.group(1))

                            # Extraer flujo si existe
                            flujo = None
                            if m_flujo:
                                proto = m_flujo.group(1).strip()
                                src = m_flujo.group(2).strip()
                                dst = m_flujo.group(3).strip()
                                ep1, ep2 = sorted([src, dst])
                                flujo = (proto, ep1, ep2)

                            # Actualizamos simultáneamente la prioridad específica y el global "Any"
                            claves_a_actualizar = [(rs_num, prio), (rs_num, 'Any')]

                            for clave_tupla in claves_a_actualizar:
                                if clave_tupla not in datos_completos[ataque_base]:
                                    datos_completos[ataque_base][clave_tupla] = {
                                        'alertas': 0,
                                        'sids': Counter(),
                                        'flujos': set(),
                                        'flujos_sids': {}  # <--- NUEVO: Diccionario para trackear SIDs por flujo
                                    }

                                stats = datos_completos[ataque_base][clave_tupla]
                                stats['alertas'] += 1
                                stats['sids'][sid] += 1

                                if flujo:
                                    stats['flujos'].add(flujo)
                                    # Registrar el SID asociado a este flujo concreto (usando un set para evitar duplicados)
                                    if flujo not in stats['flujos_sids']:
                                        stats['flujos_sids'][flujo] = set()
                                    stats['flujos_sids'][flujo].add(sid)
                                else:
                                    # Por si hay alertas sin información de flujo, las agrupamos para no perder la cuenta
                                    if None not in stats['flujos_sids']:
                                        stats['flujos_sids'][None] = set()
                                    stats['flujos_sids'][None].add(sid)

            except Exception as e:
                print(f"  [!] Error leyendo {filename}: {e}")

    # --- RELLENAR LA PLANTILLA EXCEL ---
    print(f"\n[+] Abriendo plantilla para volcar los datos...")
    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active

    # Configurar estilo de celda
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_style = Alignment(vertical='top', wrap_text=True, horizontal='left')

    # 1. Mapear las columnas maestras (Soporta números y "Any" explícito o implícito)
    mapeo_columnas = {}

    for col in range(1, ws.max_column + 1):
        valor_celda = str(ws.cell(row=1, column=col).value or "").strip()

        if valor_celda:
            # Buscar RSX
            m_rs = re.search(r'RS(\d+)', valor_celda, re.IGNORECASE)
            if m_rs:
                r = int(m_rs.group(1))

                # Buscar prioridad
                m_pri = re.search(r'Priority\s*=\s*(\d+|Any)', valor_celda, re.IGNORECASE)
                if m_pri:
                    pri_val = m_pri.group(1)
                    current_priority = int(pri_val) if pri_val.isdigit() else 'Any'
                else:
                    current_priority = 'Any'  # Si no especifica, asumimos Any

                mapeo_columnas[(r, current_priority)] = col

    # 2. Insertar los datos fila a fila (Rellenando con CEROS)
    fila_actual = 3
    for ataque in orden_list:
        # Columna A: Nombre del ataque
        celda_ataque = ws.cell(row=fila_actual, column=1, value=ataque)
        celda_ataque.border = border_thin
        celda_ataque.alignment = alignment_style

        # Rellenar cada bloque de RS + Prioridad encontrado
        for (r, p), col_base in mapeo_columnas.items():
            stats = datos_completos[ataque].get((r, p), None)

            # AHORA HAY 6 SUBCOLUMNAS (añadida la de Alertas diferentes por flujo)
            for offset in range(6):
                col_actual = col_base + offset
                cabecera_sub = str(ws.cell(row=2, column=col_actual).value).strip().lower()

                valor = 0  # Inicializamos en 0 por defecto

                if stats and stats['alertas'] > 0:
                    # Orden de comprobación estricto para que las subcadenas no hagan conflicto
                    if "diferentes por flujo" in cabecera_sub:
                        # NUEVA MÉTRICA: Suma de la cantidad de SIDs únicos que tiene cada flujo de este pcap
                        valor = sum(len(sids_unicos) for sids_unicos in stats['flujos_sids'].values())
                    elif "flujo" in cabecera_sub:
                        valor = len(stats['flujos'])
                    elif cabecera_sub == "#alertas" or cabecera_sub == "#alerts":
                        valor = stats['alertas']
                    elif cabecera_sub == "#alertas/sid" or "alertas/sid" in cabecera_sub:
                        sids_ord = sorted(stats['sids'].items(), key=lambda x: x[1], reverse=True)
                        valor = ", ".join([f"{s} ({c})" for s, c in sids_ord])
                    elif "sin repetición" in cabecera_sub or "sin repeticion" in cabecera_sub:
                        sids_ord = sorted(stats['sids'].items(), key=lambda x: x[1], reverse=True)
                        valor = ", ".join([str(s) for s, _ in sids_ord])
                    elif cabecera_sub == "#sid" or cabecera_sub == "#sids":
                        valor = len(stats['sids'])

                celda_dato = ws.cell(row=fila_actual, column=col_actual, value=valor)
                celda_dato.border = border_thin
                celda_dato.alignment = alignment_style

        fila_actual += 1

    wb.save(ruta_salida)
    print(f"✅ ¡Operación completada con éxito! Excel generado: {ruta_salida}")


if __name__ == "__main__":
    main()