import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font


def get_summary_metrics(summary_path):
    metrics = {}
    with open(summary_path, 'r', encoding='utf-8') as f:
        content = f.read()

    blocks_map = {
        "PRUNING 500": "[PRUNING 500]",
        "PRUNING 1000": "[PRUNING 1000]",
        "CONSERVATIVE PRUNING": "[CONSERVATIVE PRUNING]",
        "500+CONSERVATIVE": "[500+CONSERVATIVE]",
        "1000+CONSERVATIVE": "[1000+CONSERVATIVE]",
        "AGRESIVE PRUNING": "[AGRESIVE PRUNING]"
    }

    if "[INITIAL STATE (After Phase 1" in content:
        blocks_map["CONSERVATIVE PRUNING"] = "[INITIAL STATE (After Phase 1 - Exclusive FPs Removed)]"
        blocks_map["AGRESIVE PRUNING"] = "[FINAL STATE (Optimal G-Mean Reached)]"

    for dict_key, txt_header in blocks_map.items():
        if txt_header in content:
            blk = content.split(txt_header)[1].split('[')[0] if '[' in content.split(txt_header)[1] else \
            content.split(txt_header)[1]
            m = {}
            try:
                m['active_rules'] = int(re.search(r'Active SIDs:\s*(\d+)', blk).group(1)) if re.search(r'Active SIDs:',
                                                                                                       blk) else int(
                    re.search(r'Active Rules:\s*(\d+)', blk).group(1))

                # Extracción robusta de los SIDs eliminados directamente del TXT
                match_rem = re.search(r'Removed SIDs:\s*(.*)', blk)
                if match_rem:
                    m['removed_sids'] = match_rem.group(1).strip()
                    m['removed_count'] = len(
                        [s for s in m['removed_sids'].split(',') if s.strip() and s.strip() != 'None']) if m[
                                                                                                               'removed_sids'] != 'None' else 0
                else:
                    m['removed_sids'] = "None"
                    m['removed_count'] = 0

                m['gm'] = float(re.search(r'G-Mean:\s*([0-9.]+)', blk).group(1))

                match_tpr = re.search(r'TPR:\s*([0-9.]+)', blk)
                if match_tpr:
                    m['tpr'] = float(match_tpr.group(1)) * 100
                else:
                    m['tpr'] = float(re.search(r'Detection Rate \(PCAPs\):\s*([0-9.]+)', blk).group(1)) * 100

                m['a_tp'] = int(re.search(r'Raw Alerts \(Attacks\):\s*(\d+)', blk).group(1))
                m['a_tp_diff'] = int(re.search(r'Diff Alerts/Flow \(Attacks\):\s*(\d+)', blk).group(1))
                m['a_fp'] = int(re.search(r'Raw Alerts \(Legitimate\):\s*(\d+)', blk).group(1))
                m['a_fp_diff'] = int(re.search(r'Diff Alerts/Flow \(Legitimate\):\s*(\d+)', blk).group(1))
                m['a_tot'] = int(re.search(r'Raw Alerts \(Total\):\s*(\d+)', blk).group(1))
                m['usab'] = float(re.search(r'Usability \(FP_Alerts/TP_Alerts\):\s*([0-9.]+)', blk).group(1))
                metrics[dict_key] = m
            except Exception as e:
                pass
    return metrics


def locate_block_row(sheet, block_name):
    """Busca dinámicamente en qué fila de la plantilla empieza cada bloque."""
    for r in range(1, sheet.max_row + 1):
        val = str(sheet.cell(row=r, column=1).value).strip()
        if not val or val == 'None': continue
        if val == block_name:
            return r
    return None


def find_row(sheet, keyword, start_row, end_row):
    for r in range(start_row, end_row + 1):
        cell_val = str(sheet.cell(row=r, column=1).value).strip()
        if cell_val == keyword or cell_val.startswith(keyword):
            return r
    return None


def write_metric(sheet, keyword, start_r, end_r, col, val, align_left):
    if not start_r or not end_r: return
    r = find_row(sheet, keyword, start_r, end_r)
    if r:
        cell = sheet.cell(row=r, column=col, value=val)
        if isinstance(val, str) and ',' in val:
            cell.alignment = align_left


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)
    print("=== TFG: GENERATOR OF FINAL TABLES & GLOBAL SUMMARY ===\n")

    master_dir = "Resultados_Snort_GM"
    plantilla_path = "Resultados_Snort_GM/plantilla.xlsx"
    salida_tablas = os.path.join(master_dir, "Snort_GM_Final_Tables.xlsx")
    salida_resumen = os.path.join(master_dir, "Global_G-Mean_Snort_Summary.xlsx")

    if not os.path.exists(plantilla_path):
        print(f"[!] Error: Template '{plantilla_path}' no encontrada.")
        return

    print("[*] TAREA 1: Rellenando plantilla detallada por hoja de prioridad...")
    wb_tables = openpyxl.load_workbook(plantilla_path)
    ws_base = wb_tables.active

    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center')

    rs_cols = {}
    for col in range(2, ws_base.max_column + 1):
        for r in range(1, 5):
            val = str(ws_base.cell(row=r, column=col).value).strip()
            m = re.search(r'RS(\d+)', val, re.IGNORECASE)
            if m:
                rs_cols[f"RS{m.group(1)}"] = col
                break

    if not rs_cols:
        print("[!] ERROR: No se detectaron cabeceras RS1, RS2 en la plantilla.")
        return

    escenarios = [4, 3, 2, 1]

    blocks_in_txt = [
        "PRUNING 500", "PRUNING 1000", "CONSERVATIVE PRUNING",
        "500+CONSERVATIVE", "1000+CONSERVATIVE", "AGRESIVE PRUNING"
    ]

    global_summary_data = []
    history_by_scenario = {}

    for pri in escenarios:
        ws_new = wb_tables.copy_worksheet(ws_base)
        ws_new.title = f"Pri <= {pri}"
        print(f"\n  >> Rellenando Hoja: Pri <= {pri}")

        row_bounds = {}
        for r in range(1, ws_new.max_row + 1):
            val = str(ws_new.cell(row=r, column=1).value).strip()
            if val in blocks_in_txt:
                row_bounds[val] = r

        row_rs_abajo = None
        start_search = row_bounds.get("AGRESIVE PRUNING", 80) if row_bounds.get("AGRESIVE PRUNING") else 80
        for r in range(start_search, ws_new.max_row + 20):
            if str(ws_new.cell(row=r, column=2).value).strip() == "RS1":
                row_rs_abajo = r
                break

        for rs, col_num in rs_cols.items():
            res_folder = os.path.join(master_dir, rs, f"Pri_{pri}")
            summary_path = os.path.join(res_folder, "3_Optimization_Summary.txt")

            if not os.path.exists(summary_path):
                continue

            metrics = get_summary_metrics(summary_path)
            if not metrics: continue

            # --- RECOLECTAR DATOS PARA EL RESUMEN GLOBAL ---
            m_cons = metrics.get("CONSERVATIVE PRUNING")
            m_agr = metrics.get("AGRESIVE PRUNING")

            if m_cons and m_agr:
                rules_removed = m_cons['active_rules'] - m_agr['active_rules']
                gm_improvement = m_agr['gm'] - m_cons['gm']

                tpr_ini_dec = m_cons['tpr'] / 100.0
                tnr_ini_calc = (m_cons['gm'] ** 2) / tpr_ini_dec if tpr_ini_dec > 0 else 0.0

                tpr_fin_dec = m_agr['tpr'] / 100.0
                tnr_fin_calc = (m_agr['gm'] ** 2) / tpr_fin_dec if tpr_fin_dec > 0 else 0.0

                label_global = f"{rs} [Pri<={pri}]"

                history_path = os.path.join(res_folder, "2_GM_Optimization_Steps.csv")
                history_f2 = []
                if os.path.exists(history_path):
                    try:
                        df_hist = pd.read_csv(history_path)
                        if 'Step' in df_hist.columns and 'G-Mean' in df_hist.columns:
                            history_f2 = [(int(row['Step']), float(row['G-Mean'])) for _, row in df_hist.iterrows()]
                    except:
                        pass
                history_by_scenario[label_global] = history_f2

                global_summary_data.append({
                    'Ruleset': rs,
                    'Scenario': f"Pri <= {pri}",
                    'Initial_Active_SIDs': m_cons['active_rules'],
                    'Initial_TP_PCAPs': m_cons['a_tp'],
                    'Initial_FP_Flows': m_cons['a_fp'],
                    'Initial_TPR': round(tpr_ini_dec, 6),
                    'Initial_TNR': round(tnr_ini_calc, 6),
                    'Initial_G-Mean': round(m_cons['gm'], 6),
                    'SIDs_Removed_Optimization': rules_removed,
                    'Final_Active_SIDs': m_agr['active_rules'],
                    'Final_TP_PCAPs': m_agr['a_tp'],
                    'Final_FP_Flows': m_agr['a_fp'],
                    'Final_TPR': round(tpr_fin_dec, 6),
                    'Final_TNR': round(tnr_fin_calc, 6),
                    'Final_G-Mean': round(m_agr['gm'], 6),
                    'G-Mean_Improvement': round(gm_improvement, 6),
                    'GM_Pre500': round(metrics.get("PRUNING 500", {}).get('gm', 0.0), 6),
                    'GM_Pre1000': round(metrics.get("PRUNING 1000", {}).get('gm', 0.0), 6),
                    'GM_Post500': round(metrics.get("500+CONSERVATIVE", {}).get('gm', 0.0), 6),
                    'GM_Post1000': round(metrics.get("1000+CONSERVATIVE", {}).get('gm', 0.0), 6)
                })

            # --- VOLCAR LAS MÉTRICAS EN LA PLANTILLA DETALLADA ---
            for b_name in blocks_in_txt:
                b_start = row_bounds.get(b_name)
                if not b_start: continue

                m = metrics.get(b_name)
                if not m: continue

                # ¡Escribimos las variables extraídas directamente del TXT!
                write_metric(ws_new, "GM", b_start, b_start + 15, col_num, m['gm'], align_left)
                write_metric(ws_new, "SIDs/AttackID", b_start, b_start + 15, col_num, m['removed_sids'], align_left)
                write_metric(ws_new, "#SIDs/AttackID", b_start, b_start + 15, col_num, m['removed_count'], align_left)
                write_metric(ws_new, "#Alertas total", b_start, b_start + 15, col_num, m['a_tot'], align_left)
                write_metric(ws_new, "%Detección_pcaps", b_start, b_start + 15, col_num, f"{m['tpr']:.2f}%", align_left)
                write_metric(ws_new, "#Alertas_TP", b_start, b_start + 15, col_num, m['a_tp'], align_left)
                write_metric(ws_new, "#Alertas TP (Diff", b_start, b_start + 15, col_num, m['a_tp_diff'], align_left)
                write_metric(ws_new, "#Alertas_FP", b_start, b_start + 15, col_num, m['a_fp'], align_left)
                write_metric(ws_new, "#Alertas FP (Diff", b_start, b_start + 15, col_num, m['a_fp_diff'], align_left)
                write_metric(ws_new, "Usabilidad", b_start, b_start + 15, col_num, m['usab'], align_left)

            # --- ESCRIBIR HISTORIAL INFERIOR ---
            if row_rs_abajo:
                for c in range(1, ws_new.max_column + 1):
                    if str(ws_new.cell(row=row_rs_abajo, column=c).value).strip() == rs:
                        row_hist = row_rs_abajo + 2
                        for h_step in history_f2:
                            c1 = ws_new.cell(row=row_hist, column=c, value=h_step[0])
                            c1.border, c1.alignment = border_thin, align_center

                            c2 = ws_new.cell(row=row_hist, column=c + 1, value=h_step[1])
                            c2.border, c2.alignment = border_thin, align_center
                            row_hist += 1
                        break

    wb_tables.remove(ws_base)
    wb_tables.save(salida_tablas)
    print(f"\n  -> ✅ Guardado fichero detallado: {salida_tablas}")

    # =========================================================================
    # PARTE 2: CREAR EL FICHERO MAESTRO GLOBAL (RESUMEN + HISTORIALES)
    # =========================================================================
    if global_summary_data:
        print("\n[*] TAREA 2: Generando el resumen global en fichero aparte...")
        df_summary = pd.DataFrame(global_summary_data)

        def rs_val(x):
            m = re.search(r'RS(\d+)', x)
            return int(m.group(1)) if m else 0

        def pri_val(x):
            m = re.search(r'Pri <= (\d+)', x)
            return -int(m.group(1)) if m else 0

        df_summary['rs_order'] = df_summary['Ruleset'].apply(rs_val)
        df_summary['pri_order'] = df_summary['Scenario'].apply(pri_val)
        df_summary.sort_values(by=['rs_order', 'pri_order'], inplace=True)
        df_summary.drop(columns=['rs_order', 'pri_order'], inplace=True)

        font_bold = Font(bold=True)
        wb_global = openpyxl.Workbook()
        ws_global = wb_global.active
        ws_global.title = "G-Mean_Evolution"
        ws_global.sheet_view.showGridLines = True

        headers = list(df_summary.columns)
        ws_global.append(headers)
        for cell in ws_global[1]:
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border_thin

        for _, row in df_summary.iterrows():
            ws_global.append(list(row))
            for cell in ws_global[ws_global.max_row]:
                cell.border = border_thin
                cell.alignment = align_center

        start_history_row = ws_global.max_row + 4

        def sort_label(x):
            rs_m = re.search(r'RS(\d+)', x)
            pri_m = re.search(r'Pri<=(\d+)', x)
            return (int(rs_m.group(1)) if rs_m else 0, -int(pri_m.group(1)) if pri_m else 0)

        sorted_labels = sorted(history_by_scenario.keys(), key=sort_label)

        col_idx = 1
        for label in sorted_labels:
            steps = history_by_scenario[label]
            if not steps: continue

            ws_global.merge_cells(start_row=start_history_row, start_column=col_idx, end_row=start_history_row,
                                  end_column=col_idx + 1)
            cell = ws_global.cell(row=start_history_row, column=col_idx, value=label)
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border_thin
            ws_global.cell(row=start_history_row, column=col_idx + 1).border = border_thin

            c1 = ws_global.cell(row=start_history_row + 1, column=col_idx, value="Nº SIDs eliminados")
            c2 = ws_global.cell(row=start_history_row + 1, column=col_idx + 1, value="GM")
            c1.font, c1.alignment, c1.border = font_bold, align_center, border_thin
            c2.font, c2.alignment, c2.border = font_bold, align_center, border_thin

            current_data_row = start_history_row + 2
            for step_num, gm_val in steps:
                c_step = ws_global.cell(row=current_data_row, column=col_idx, value=step_num)
                c_gm = ws_global.cell(row=current_data_row, column=col_idx + 1, value=gm_val)
                c_step.border, c_step.alignment = border_thin, align_center
                c_gm.border, c_gm.alignment = border_thin, align_center
                current_data_row += 1

            col_idx += 3

        for col in ws_global.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_global.column_dimensions[col_letter].width = max_len + 3

        wb_global.save(salida_resumen)
        print(f"  -> ✅ Guardado fichero independiente: {salida_resumen}")

    print("\n✅ ¡Todos los procesos completados correctamente!")


if __name__ == "__main__":
    main()