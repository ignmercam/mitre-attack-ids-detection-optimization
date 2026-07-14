import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# --- CONSTANTES GLOBALES ---
TOTAL_LEG_FLOWS = 78694206

FG_SEV_MAP = {'info': 1, 'low': 2, 'medium': 3, 'high': 4, 'critical': 5}
FG_RISK_MAP = {'information': 1, 'low': 2, 'medium': 3, 'high': 4, 'elevated': 5}


def parse_snort_raw():
    """Escanea las carpetas RSX/Legitimo y extrae el tráfico legítimo agrupado por Ruleset y Flujo."""
    # snort_data[rs_num][flow_id] = [(sid, priority), (sid, priority)...]
    snort_data = defaultdict(lambda: defaultdict(list))

    regex_sid = re.compile(r'\[\*\*\]\s+\[\d+:(\d+):\d+\]')
    regex_prio = re.compile(r'\[Priority:\s*(\d+)\]', re.IGNORECASE)
    regex_flujo = re.compile(r'\{(.*?)\}\s+(\S+)\s+->\s+(\S+)')

    carpetas = [d for d in os.listdir('.') if os.path.isdir(d) and d.startswith('RS')]

    for d in carpetas:
        m_rs = re.search(r'RS(\d+)', d)
        if not m_rs: continue
        rs_num = int(m_rs.group(1))

        leg_dir = os.path.join(d, "Legitimo")
        if not os.path.exists(leg_dir): continue

        for root, _, files in os.walk(leg_dir):
            for f in files:
                if f.endswith('.log') or f.endswith('.txt'):
                    pcap_name = f.replace('.log', '').replace('.txt', '')
                    with open(os.path.join(root, f), 'r', errors='ignore') as file:
                        for line in file:
                            m_sid = regex_sid.search(line)
                            m_prio = regex_prio.search(line)
                            m_flujo = regex_flujo.search(line)

                            if m_sid and m_prio:
                                sid = int(m_sid.group(1))
                                pri = int(m_prio.group(1))

                                if m_flujo:
                                    proto = m_flujo.group(1).strip()
                                    ep1, ep2 = sorted([m_flujo.group(2).strip(), m_flujo.group(3).strip()])
                                    flujo_id = f"{pcap_name}__{proto} {ep1} <-> {ep2}"
                                else:
                                    flujo_id = f"{pcap_name}__{line.strip()}"

                                snort_data[rs_num][flujo_id].append((sid, pri))

    return snort_data


def parse_fw_raw(engine_type):
    """Escanea las carpetas FG/PA/Legitimo y extrae el tráfico agrupado por engine y Session ID global."""
    # fw_data[session_id] = [(ips_val, ips_sev, app_val, app_sev), ...]
    fw_data = defaultdict(list)
    base_path = os.path.join(engine_type, "Legitimo")

    if not os.path.exists(base_path):
        return fw_data

    if engine_type == 'FG':
        re_sess = re.compile(r'sessionid=[\'"]?(\d+)', re.IGNORECASE)
        re_ips_id = re.compile(r'attackid=[\'"]?(\d+)', re.IGNORECASE)
        re_ips_sev = re.compile(r'severity=[\'"]?([a-zA-Z]+)', re.IGNORECASE)
        re_app_id = re.compile(r'appid=[\'"]?(\d+)', re.IGNORECASE)
        re_app_sev = re.compile(r'apprisk=[\'"]?([a-zA-Z]+)', re.IGNORECASE)
    else:
        re_sess = re.compile(r'sessionid=[\'"]?(\d+)', re.IGNORECASE)
        re_ips_id = re.compile(r'threat_id=[\'"]?([^"\'\s,]+)', re.IGNORECASE)
        re_ips_sev = re.compile(r'severity_number=[\'"]?(\d+)', re.IGNORECASE)
        re_app_id = re.compile(r'(?:appid|app)=[\'"]?([^"\'\s,]+)', re.IGNORECASE)
        re_app_sev = re.compile(r'risk_of_app=[\'"]?(\d+)', re.IGNORECASE)

    for root, dirs, files in os.walk(base_path):
        for f_name in files:
            if f_name.endswith('.log') or f_name.endswith('.txt'):
                with open(os.path.join(root, f_name), 'r', errors='ignore') as file:
                    for line in file:
                        if ('type="traffic"' in line) or ('type=traffic' in line) or (
                                'TRAFFIC' in line and engine_type == 'PA'):
                            continue

                        # Extracción global (ID único por sesión sin importar fichero)
                        match_session = re_sess.search(line)
                        session_id = match_session.group(1) if match_session else None
                        if not session_id:
                            continue

                        ips_val, ips_level = None, 0
                        m_ips_id = re_ips_id.search(line)
                        m_ips_sev = re_ips_sev.search(line)
                        if m_ips_id and m_ips_sev:
                            ips_raw = m_ips_id.group(1)
                            ips_val = re.search(r'\((\d+)\)', ips_raw).group(1) if (
                                        engine_type == 'PA' and '(' in ips_raw) else ips_raw
                            sev_raw = m_ips_sev.group(1).lower()
                            ips_level = FG_SEV_MAP.get(sev_raw, 0) if engine_type == 'FG' else int(sev_raw)

                        app_val, app_level = None, 0
                        m_app_id = re_app_id.search(line)
                        m_app_sev = re_app_sev.search(line)
                        if m_app_id and m_app_sev:
                            app_val = m_app_id.group(1)
                            sev_raw = m_app_sev.group(1).lower()
                            app_level = FG_RISK_MAP.get(sev_raw, 0) if engine_type == 'FG' else int(sev_raw)

                        if ips_val or app_val:
                            fw_data[session_id].append((ips_val, ips_level, app_val, app_level))

    return fw_data


def format_percentage(value):
    return f"{value * 100:.4f}%"


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)
    print("=== TFG: GENERADOR DE TABLAS BASE DE LEGÍTIMO (SIN PRUNING) ===\n")

    print("[*] Escaneando Logs RAW de Snort (RS1-RS9)...")
    snort_data = parse_snort_raw()

    print("[*] Escaneando Logs RAW de FortiGate y Palo Alto...")
    fg_data = parse_fw_raw('FG')
    pa_data = parse_fw_raw('PA')

    print("[+] Logs cargados. Procesando escenarios...")

    # Tupla: (Nombre Pestaña, Max_Pri_Snort, Min_Sev_FG_PA)
    # None en Snort significa Caso Any (entran todas)
    escenarios = [
        ("Caso Any", None, 1),
        ("Snort<=4_FGPA>=1", 4, 1),
        ("Snort<=3_FGPA>=2", 3, 2),
        ("Snort<=2_FGPA>=3", 2, 3),
        ("Snort<=1_FGPA>=4", 1, 4),
        ("FGPA>=5", -1, 5)  # Snort -1 = No aplica / N/A
    ]

    rulesets = [f"RS{i}" for i in range(1, 10)]
    metricas = [
        "#Alerts",
        "#Flow-correlated Alerts",
        "#SIDs, #AttackID o #ThreatID",
        "#AppID",
        "#Flows_FP",
        "#Flows_TN",
        "%Flows_FP",
        "%Flows_TN"
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for titulo_hoja, max_pri, min_sev in escenarios:
        print(f"  -> Generando hoja: {titulo_hoja}")
        ws = wb.create_sheet(title=titulo_hoja)

        # 1. Cabeceras
        ws.cell(row=1, column=1, value="Metric").font = font_bold
        ws.cell(row=1, column=1).alignment = align_center
        ws.cell(row=1, column=1).border = border_thin
        ws.cell(row=1, column=1).fill = fill_header

        col_idx = 2
        for rs in rulesets:
            c = ws.cell(row=1, column=col_idx, value=rs)
            c.font, c.alignment, c.border, c.fill = font_bold, align_center, border_thin, fill_header
            col_idx += 1

        headers_fw = [
            f"FG [IPS severity >={min_sev} (Any)]",
            f"FG [IPS severity >={min_sev} (Any)] || [App apprisk >= {min_sev} (Any)]",
            f"PA [IPS severity >={min_sev} (Any)]",
            f"PA [IPS severity_number >={min_sev} (Any)] || [App risk_of_app >= {min_sev} (Any)]"
        ]

        for h in headers_fw:
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font, c.alignment, c.border, c.fill = font_bold, align_center, border_thin, fill_header
            col_idx += 1

        for r_idx, met in enumerate(metricas, start=2):
            c = ws.cell(row=r_idx, column=1, value=met)
            c.font, c.border, c.alignment = font_bold, border_thin, align_left

        # ==========================================
        # 2. CÁLCULO SNORT (Columnas B a J)
        # ==========================================
        for c_idx, rs_name in enumerate(rulesets, start=2):
            rs_num = int(rs_name.replace('RS', ''))

            if max_pri == -1:  # Caso N/A
                for r_idx in range(2, 10):
                    ws.cell(row=r_idx, column=c_idx, value="N/A").border = border_thin
                    ws.cell(row=r_idx, column=c_idx).alignment = align_center
                continue

            alerts = 0
            alerts_diff = 0
            unique_sids = set()
            flows_fp = 0

            # Iterar sobre todos los flujos de este RS
            for flow_id, events in snort_data[rs_num].items():
                # Filtrar eventos que cumplen la prioridad
                valid_events = [sid for sid, pri in events if max_pri is None or pri <= max_pri]

                if valid_events:
                    flows_fp += 1
                    alerts += len(valid_events)
                    alerts_diff += len(set(valid_events))  # Alertas diferentes por flujo
                    unique_sids.update(valid_events)

            flows_tn = TOTAL_LEG_FLOWS - flows_fp

            data_col = [
                alerts,
                alerts_diff,
                len(unique_sids),
                "N/A",
                flows_fp,
                flows_tn,
                format_percentage(flows_fp / TOTAL_LEG_FLOWS),
                format_percentage(flows_tn / TOTAL_LEG_FLOWS)
            ]
            for r_idx, val in enumerate(data_col, start=2):
                ws.cell(row=r_idx, column=c_idx, value=val).border = border_thin
                ws.cell(row=r_idx, column=c_idx).alignment = align_center

        # ==========================================
        # 3. CÁLCULO FG y PA (Columnas K a N)
        # ==========================================
        fw_col_start = 11
        for i, engine in enumerate(['FG', 'PA']):
            fw_events_dict = fg_data if engine == 'FG' else pa_data

            # --- Columna IPS Only ---
            alerts_ips = 0
            alerts_diff_ips = 0
            unique_ids_ips = set()
            flows_fp_ips = 0

            for session_id, events in fw_events_dict.items():
                flow_unique_items = set()
                # Filtrar por validez de IPS y Severidad
                valid_events = [ips_v for ips_v, ips_l, _, _ in events if ips_v and ips_l >= min_sev]

                if valid_events:
                    flows_fp_ips += 1
                    alerts_ips += len(valid_events)
                    alerts_diff_ips += len(set(valid_events))
                    unique_ids_ips.update(valid_events)

            flows_tn_ips = TOTAL_LEG_FLOWS - flows_fp_ips

            col_ips = fw_col_start + (i * 2)
            data_ips = [
                alerts_ips,
                alerts_diff_ips,
                len(unique_ids_ips),
                "N/A",
                flows_fp_ips,
                flows_tn_ips,
                format_percentage(flows_fp_ips / TOTAL_LEG_FLOWS),
                format_percentage(flows_tn_ips / TOTAL_LEG_FLOWS)
            ]

            for r_idx, val in enumerate(data_ips, start=2):
                ws.cell(row=r_idx, column=col_ips, value=val).border = border_thin
                ws.cell(row=r_idx, column=col_ips).alignment = align_center

            # --- Columna IPS || APP ---
            alerts_comp = 0
            alerts_diff_comp = 0
            unique_ids_comp = set()
            unique_apps_comp = set()
            flows_fp_comp = 0

            for session_id, events in fw_events_dict.items():
                flow_triggered = False
                flow_unique_items = set()

                for ips_val, ips_sev, app_val, app_sev in events:
                    valid_ips = (ips_val and ips_sev >= min_sev)
                    valid_app = (app_val and app_sev >= min_sev)

                    if valid_ips or valid_app:
                        flow_triggered = True
                        if valid_ips:
                            alerts_comp += 1
                            flow_unique_items.add(ips_val)
                            unique_ids_comp.add(ips_val)
                        if valid_app:
                            alerts_comp += 1
                            flow_unique_items.add(app_val)
                            unique_apps_comp.add(app_val)

                if flow_triggered:
                    flows_fp_comp += 1
                    alerts_diff_comp += len(flow_unique_items)

            flows_tn_comp = TOTAL_LEG_FLOWS - flows_fp_comp

            col_comp = fw_col_start + (i * 2) + 1
            data_comp = [
                alerts_comp,
                alerts_diff_comp,
                len(unique_ids_comp),
                len(unique_apps_comp),
                flows_fp_comp,
                flows_tn_comp,
                format_percentage(flows_fp_comp / TOTAL_LEG_FLOWS),
                format_percentage(flows_tn_comp / TOTAL_LEG_FLOWS)
            ]

            for r_idx, val in enumerate(data_comp, start=2):
                ws.cell(row=r_idx, column=col_comp, value=val).border = border_thin
                ws.cell(row=r_idx, column=col_comp).alignment = align_center

        # Ajuste de ancho de columnas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    output_filename = "Resultados_Legitimo_Base.xlsx"
    wb.save(output_filename)
    print(f"\n✅ ¡PROCESO COMPLETADO! Tablas generadas en: {output_filename}")


if __name__ == '__main__':
    main()