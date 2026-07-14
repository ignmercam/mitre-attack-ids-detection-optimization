import os
import re
import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter

# --- MAPAS DE SEVERIDAD/RIESGO A NIVEL NUMÉRICO (1-5) -----------------
FG_SEV_MAP = {'info': 1, 'low': 2, 'medium': 3, 'high': 4, 'critical': 5}
FG_RISK_MAP = {'information': 1, 'low': 2, 'medium': 3, 'high': 4, 'elevated': 5}

# --- REGEX POR MOTOR (Idénticas a optimo_FG-PA_GM_v6.py) --------------
ENGINE_REGEX = {
    'FG': {
        'traffic_line': re.compile(r'type=["\']?traffic["\']?', re.IGNORECASE),
        'session': re.compile(r'sessionid=[\'"]?(\d+)', re.IGNORECASE),
        'ips_id': re.compile(r'attackid=[\'"]?(\d+)', re.IGNORECASE),
        'ips_sev': re.compile(r'severity=[\'"]?([a-zA-Z]+)', re.IGNORECASE),
        'app_id': re.compile(r'appid=[\'"]?(\d+)', re.IGNORECASE),
        'app_sev': re.compile(r'apprisk=[\'"]?([a-zA-Z]+)', re.IGNORECASE),
    },
    'PA': {
        'traffic_line': re.compile(r'log_type=["\']?TRAFFIC["\']?', re.IGNORECASE),
        'session': re.compile(r'sessionid=[\'"]?(\d+)', re.IGNORECASE),
        'ips_id': re.compile(r'threat_id=[\'"]?([^"\'\s,]+)', re.IGNORECASE),
        'ips_sev': re.compile(r'severity_number=[\'"]?(\d+)', re.IGNORECASE),
        'app_id': re.compile(r'(?:appid|app)=[\'"]?([^"\'\s,]+)', re.IGNORECASE),
        'app_sev': re.compile(r'risk_of_app=[\'"]?(\d+)', re.IGNORECASE),
    }
}


def extract_pa_threat_id(raw_id):
    m = re.search(r'\((\d+)\)', raw_id)
    if m:
        return m.group(1)
    m2 = re.search(r'(\d+)', raw_id)
    if m2:
        return m2.group(1)
    return raw_id


def parse_line(line, engine, regexes):
    if regexes['traffic_line'].search(line):
        return None

    m_sess = regexes['session'].search(line)
    session_id = m_sess.group(1) if m_sess else None

    m_ips_id = regexes['ips_id'].search(line)
    m_ips_sev = regexes['ips_sev'].search(line)
    ips_id, ips_level = None, 0
    if m_ips_id and m_ips_sev:
        raw_id = m_ips_id.group(1)
        ips_id = extract_pa_threat_id(raw_id) if engine == 'PA' else raw_id
        sev_raw = m_ips_sev.group(1).lower()
        ips_level = FG_SEV_MAP.get(sev_raw, 0) if engine == 'FG' else (
            int(sev_raw) if sev_raw.isdigit() else 0)
        if ips_level == 0:
            ips_id = None

    m_app_id = regexes['app_id'].search(line)
    m_app_sev = regexes['app_sev'].search(line)
    app_id, app_level = None, 0
    if m_app_id and m_app_sev:
        app_id = m_app_id.group(1)
        sev_raw = m_app_sev.group(1).lower()
        app_level = FG_RISK_MAP.get(sev_raw, 0) if engine == 'FG' else (
            int(sev_raw) if sev_raw.isdigit() else 0)
        if app_level == 0:
            app_id = None

    if ips_id is None and app_id is None:
        return None

    return session_id, ips_id, ips_level, app_id, app_level


def read_engine_folder(base_path, engine, regexes):
    atk_pcap_lines = defaultdict(list)
    leg_lines = []

    ruta_atk = os.path.join(base_path, "Ataques")
    if os.path.exists(ruta_atk):
        for fname in os.listdir(ruta_atk):
            if not (fname.endswith('.log') or fname.endswith('.txt')):
                continue
            pcap_name = fname.replace('.log', '').replace('.txt', '')
            with open(os.path.join(ruta_atk, fname), 'r', errors='ignore') as f:
                for line in f:
                    parsed = parse_line(line, engine, regexes)
                    if parsed is None:
                        continue
                    session_id, ips_id, ips_level, app_id, app_level = parsed
                    flow_key = f"{pcap_name}__sess_{session_id}" if session_id else f"{pcap_name}__NO_SESSION"
                    atk_pcap_lines[pcap_name].append((flow_key, ips_id, ips_level, app_id, app_level))

    ruta_leg = os.path.join(base_path, "Legitimo")
    if os.path.exists(ruta_leg):
        for root, dirs, files in os.walk(ruta_leg):
            for fname in files:
                if not (fname.endswith('.log') or fname.endswith('.txt')):
                    continue
                with open(os.path.join(root, fname), 'r', errors='ignore') as f:
                    for line in f:
                        parsed = parse_line(line, engine, regexes)
                        if parsed is None:
                            continue
                        session_id, ips_id, ips_level, app_id, app_level = parsed
                        flow_key = f"{fname}__sess_{session_id}" if session_id else f"{fname}__NO_SESSION"
                        leg_lines.append((flow_key, ips_id, ips_level, app_id, app_level))

    return atk_pcap_lines, leg_lines


def build_scenario_structures(atk_pcap_lines, leg_lines, scenario, track):
    def filtro(ips_id, ips_level, app_id, app_level):
        v_i = ips_id if (ips_id and ips_level >= scenario) else None
        v_a = app_id if (app_id and app_level >= scenario) else None
        if track == 'IPS':
            return v_i, None
        else:
            return v_i, v_a

    atk_data_pcap = defaultdict(set)
    atk_data_flow = defaultdict(set)
    atk_alert_counts = Counter()
    atk_diff_seen = defaultdict(set)
    atk_line_co_occur = Counter()

    for pcap, lineas in atk_pcap_lines.items():
        for flow_key, ips_id, ips_level, app_id, app_level in lineas:
            v_i, v_a = filtro(ips_id, ips_level, app_id, app_level)
            is_dummy = flow_key.endswith("__NO_SESSION")

            if v_i is not None and v_a is not None:
                atk_line_co_occur[tuple(sorted((str(v_i), str(v_a))))] += 1

            for v in (v_i, v_a):
                if v is None:
                    continue
                atk_data_pcap[pcap].add(v)
                atk_alert_counts[v] += 1
                if not is_dummy:
                    atk_data_flow[flow_key].add(v)
                    atk_diff_seen[flow_key].add(v)

    atk_diff_counts = Counter()
    for ids in atk_diff_seen.values():
        for v in ids:
            atk_diff_counts[v] += 1

    leg_data_flow = defaultdict(set)
    leg_alert_counts = Counter()
    leg_diff_seen = defaultdict(set)
    leg_line_co_occur = Counter()

    for flow_key, ips_id, ips_level, app_id, app_level in leg_lines:
        v_i, v_a = filtro(ips_id, ips_level, app_id, app_level)
        is_dummy = flow_key.endswith("__NO_SESSION")

        if v_i is not None and v_a is not None:
            leg_line_co_occur[tuple(sorted((str(v_i), str(v_a))))] += 1

        for v in (v_i, v_a):
            if v is None:
                continue
            leg_alert_counts[v] += 1
            if not is_dummy:
                leg_data_flow[flow_key].add(v)
                leg_diff_seen[flow_key].add(v)

    leg_diff_counts = Counter()
    for ids in leg_diff_seen.values():
        for v in ids:
            leg_diff_counts[v] += 1

    return (atk_data_pcap, leg_data_flow, atk_alert_counts, leg_alert_counts,
            atk_diff_counts, leg_diff_counts, atk_line_co_occur, leg_line_co_occur)


def calc_metrics(test_ids, atk_data_pcap, leg_data_flow, total_atk_pcaps, total_leg_flows,
                 atk_alert_counts, leg_alert_counts, atk_diff_counts, leg_diff_counts,
                 atk_line_co_occur=None, leg_line_co_occur=None):
    tp = sum(1 for pcap_ids in atk_data_pcap.values() if not pcap_ids.isdisjoint(test_ids))
    tpr = tp / total_atk_pcaps if total_atk_pcaps > 0 else 0.0

    fp = sum(1 for flow_ids in leg_data_flow.values() if not flow_ids.isdisjoint(test_ids))
    tn = total_leg_flows - fp
    tnr = tn / total_leg_flows if total_leg_flows > 0 else 0.0

    gm = math.sqrt(tpr * tnr)

    alerts_tp = sum(atk_alert_counts[i] for i in test_ids)
    alerts_fp = sum(leg_alert_counts[i] for i in test_ids)

    if atk_line_co_occur:
        for (id1, id2), overlap_count in atk_line_co_occur.items():
            if id1 in test_ids and id2 in test_ids:
                alerts_tp -= overlap_count

    if leg_line_co_occur:
        for (id1, id2), overlap_count in leg_line_co_occur.items():
            if id1 in test_ids and id2 in test_ids:
                alerts_fp -= overlap_count

    alerts_tp_diff = sum(atk_diff_counts[i] for i in test_ids)
    alerts_fp_diff = sum(leg_diff_counts[i] for i in test_ids)

    return tpr, gm, alerts_tp, alerts_tp_diff, alerts_fp, alerts_fp_diff


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)
    print("=== TFG: CONSERVATIVE GENERALIZED PRUNING - FG/PA (BASED ON GM v6) ===\n")

    TOTAL_ATK_PCAPS = 97
    TOTAL_LEG_FLOWS = 1179571

    print("[*] Cargando Logs Originales de FG y PA en memoria...")
    logs_motores = {}
    for engine in ['FG', 'PA']:
        if os.path.exists(engine):
            logs_motores[engine] = read_engine_folder(engine, engine, ENGINE_REGEX[engine])

    if not logs_motores:
        print("[!] No se detectaron carpetas de FG o PA. Saliendo...")
        return

    escenarios_severidad = [1, 2, 3, 4, 5]
    columnas_motores = ['FG IPS', 'PA IPS', 'FG APP', 'PA APP']

    resultados_globales = {sev: {} for sev in escenarios_severidad}

    for sev_limit in escenarios_severidad:
        print(f"[*] Evaluando Escenario: Severidad Base >= {sev_limit}")
        for engine in ['FG', 'PA']:
            if engine not in logs_motores: continue
            atk_pcap_lines, leg_lines = logs_motores[engine]

            for track in ['IPS', 'APP']:
                (atk_data_pcap, leg_data_flow, atk_alert_counts, leg_alert_counts,
                 atk_diff_counts, leg_diff_counts,
                 atk_line_co_occur, leg_line_co_occur) = build_scenario_structures(
                    atk_pcap_lines, leg_lines, sev_limit, track)

                sids_ataque = set(atk_alert_counts.keys())
                sids_legitimo = set(leg_alert_counts.keys())
                sids_activos = sids_ataque.union(sids_legitimo)

                sids_cons_gen = sids_activos - sids_legitimo
                removed_sids = sids_legitimo

                tpr, gm, a_tp, a_tp_d, a_fp, a_fp_d = calc_metrics(
                    sids_cons_gen, atk_data_pcap, leg_data_flow, TOTAL_ATK_PCAPS, TOTAL_LEG_FLOWS,
                    atk_alert_counts, leg_alert_counts, atk_diff_counts, leg_diff_counts,
                    atk_line_co_occur, leg_line_co_occur
                )

                sorted_removed = sorted(list(removed_sids),
                                        key=lambda x: (0, int(str(x))) if str(x).isdigit() else (1, str(x)))
                removed_str = ", ".join(str(s) for s in sorted_removed) if sorted_removed else "None"
                removed_count = len(sorted_removed)
                a_tot = a_tp + a_fp
                usab = round(a_fp / a_tp, 4) if a_tp > 0 else 0.0

                col_name = f"{engine} {track}"
                resultados_globales[sev_limit][col_name] = {
                    'gm': gm,
                    'removed_sids': removed_str,
                    'removed_count': removed_count,
                    'a_tot': a_tot,
                    'tpr': f"{tpr * 100:.2f}%",
                    'a_tp': a_tp,
                    'a_tp_diff': a_tp_d,
                    'a_fp': a_fp,
                    'a_fp_diff': a_fp_d,
                    'usab': usab
                }

    print("\n[*] Generando archivo Excel con diseño profesional...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_title = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_metric_header = Font(name='Segoe UI', size=11, bold=True, color='1F497D')
    font_bold = Font(name='Segoe UI', size=10, bold=True)
    font_regular = Font(name='Segoe UI', size=10)

    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    fill_sub_header = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

    border_thin = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
                         top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    metrics_layout = [
        ("GM", 'gm'),
        ("SIDs/AttackID/ThreatID o AppID eliminados", 'removed_sids'),
        ("#SIDs/AttackID/ThreatID o AppID eliminados", 'removed_count'),
        ("#Alertas total", 'a_tot'),
        ("%Detección_pcaps", 'tpr'),
        ("#Alertas_TP", 'a_tp'),
        ("#Alertas TP (Different/Flow)", 'a_tp_diff'),
        ("#Alertas_FP", 'a_fp'),
        ("#Alertas FP (Different/Flow)", 'a_fp_diff'),
        ("Usabilidad", 'usab')
    ]

    for sev in escenarios_severidad:
        ws = wb.create_sheet(title=f"Sev >= {sev}")
        ws.sheet_view.showGridLines = True

        ws.cell(row=1, column=1, value="Metric").font = font_title
        ws.cell(row=1, column=1).fill = fill_header
        ws.cell(row=1, column=1).alignment = align_center
        ws.cell(row=1, column=1).border = border_thin

        for col_idx, motor_name in enumerate(columnas_motores, start=2):
            cell = ws.cell(row=1, column=col_idx, value=motor_name)
            cell.font, cell.fill, cell.alignment, cell.border = font_title, fill_header, align_center, border_thin

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas_motores) + 1)
        title_cell = ws.cell(row=2, column=1, value="Conservative Generalized Pruning (0% FP Tolerance)")
        title_cell.font, title_cell.fill, title_cell.alignment, title_cell.border = font_metric_header, fill_sub_header, Alignment(
            horizontal='left', vertical='center', indent=1), border_thin

        for col_idx in range(1, len(columnas_motores) + 2):
            ws.cell(row=2, column=col_idx).border = border_thin

        for row_offset, (label, key) in enumerate(metrics_layout, start=3):
            lbl_cell = ws.cell(row=row_offset, column=1, value=label)
            lbl_cell.font, lbl_cell.border, lbl_cell.alignment = font_bold, border_thin, align_left

            for col_idx, motor_name in enumerate(columnas_motores, start=2):
                val = resultados_globales[sev].get(motor_name, {}).get(key, "")
                if isinstance(val, float):
                    val = round(val, 6)

                val_cell = ws.cell(row=row_offset, column=col_idx, value=val)
                val_cell.font, val_cell.border = font_regular, border_thin
                val_cell.alignment = align_left if key == 'removed_sids' else align_center

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 2: continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        ws.column_dimensions['A'].width = 42

    output_filename = "FG-PA_Conservative_Generalized_Pruning.xlsx"
    wb.save(output_filename)
    print(f"\n✅ EXCEL PROCESADO Y GUARDADO CON ÉXITO: {output_filename}")


if __name__ == "__main__":
    main()