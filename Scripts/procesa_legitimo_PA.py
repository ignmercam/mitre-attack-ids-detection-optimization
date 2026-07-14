import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEV_PA = ['1', '2', '3', '4', '5']
RISK_PA = ['1', '2', '3', '4', '5']


def procesar_logs_legitimo_pa(ruta_legitimo):
    """Lee todos los logs de tráfico legítimo y los aglutina en una única macro-fila."""
    sessions_any = set()
    sessions_threat = set()
    sessions_app = set()

    total_alerts = 0
    unique_threat_ids = set()
    unique_apps = set()

    threat_by_sev = {s: {"alerts": 0, "ids": set(), "sessions": set(), "session_ids": {}} for s in SEV_PA}
    app_by_risk = {r: {"alerts": 0, "ids": set(), "sessions": set(), "session_ids": {}} for r in RISK_PA}

    global_session_events = {}
    archivos_procesados = 0

    for root, dirs, files in os.walk(ruta_legitimo):
        for f_name in files:
            if f_name.endswith('.log') or f_name.endswith('.txt'):
                archivos_procesados += 1
                ruta_archivo = os.path.join(root, f_name)

                with open(ruta_archivo, 'r', errors='ignore') as f:
                    for line in f:
                        if re.search(r'log_type=["\']?TRAFFIC["\']?', line, re.IGNORECASE):
                            continue
                        match_session = re.search(r'sessionid=[\'"]?(\d+)', line, re.IGNORECASE)
                        match_threat = re.search(r'threat_id=[\'"]?([^"\'\s,]+)', line, re.IGNORECASE)
                        match_sev = re.search(r'severity_number=[\'"]?(\d+)', line, re.IGNORECASE)
                        match_app = re.search(r'(?:appid|app)=[\'"]?([^"\'\s,]+)', line, re.IGNORECASE)
                        match_risk = re.search(r'risk_of_app=[\'"]?(\d+)', line, re.IGNORECASE)

                        # Unificar el id del fichero para que flujos de dias distintos no colisionen
                        session_id = f"{f_name}_{match_session.group(1)}" if match_session else None

                        t_id = None
                        if match_threat:
                            t_id_raw = match_threat.group(1)
                            id_match = re.search(r'\((\d+)\)', t_id_raw)
                            t_id = id_match.group(1) if id_match else re.search(r'(\d+)', t_id_raw).group(
                                1) if re.search(r'(\d+)', t_id_raw) else t_id_raw

                        app_name = match_app.group(1) if match_app else None

                        # --- CONTEO GLOBAL UNIFICADO ---
                        if t_id or app_name:
                            total_alerts += 1
                            if session_id:
                                sessions_any.add(session_id)
                                if session_id not in global_session_events:
                                    global_session_events[session_id] = set()
                                global_session_events[session_id].add((t_id, app_name))

                        # --- PROCESADO IPS ---
                        if match_threat:
                            unique_threat_ids.add(t_id)
                            if session_id:
                                sessions_threat.add(session_id)
                            if match_sev:
                                sev = match_sev.group(1)
                                if sev in threat_by_sev:
                                    threat_by_sev[sev]["alerts"] += 1
                                    threat_by_sev[sev]["ids"].add(t_id)
                                    if session_id:
                                        threat_by_sev[sev]["sessions"].add(session_id)
                                        if session_id not in threat_by_sev[sev]["session_ids"]:
                                            threat_by_sev[sev]["session_ids"][session_id] = set()
                                        threat_by_sev[sev]["session_ids"][session_id].add(t_id)

                        # --- PROCESADO APP ---
                        if match_app:
                            unique_apps.add(app_name)
                            if session_id:
                                sessions_app.add(session_id)
                            if match_risk:
                                risk = match_risk.group(1)
                                if risk in app_by_risk:
                                    app_by_risk[risk]["alerts"] += 1
                                    app_by_risk[risk]["ids"].add(app_name)
                                    if session_id:
                                        app_by_risk[risk]["sessions"].add(session_id)
                                        if session_id not in app_by_risk[risk]["session_ids"]:
                                            app_by_risk[risk]["session_ids"][session_id] = set()
                                        app_by_risk[risk]["session_ids"][session_id].add(app_name)

    print(f"  -> Se han procesado {archivos_procesados} archivos de tráfico legítimo.")

    global_diff_alerts = sum(len(s) for s in global_session_events.values())

    # ----------------------------------------------------
    # CONSTRUCCIÓN DE LA FILA SECUENCIAL (69 COLUMNAS)
    # ----------------------------------------------------
    res = [
        total_alerts,
        global_diff_alerts,
        len(unique_threat_ids),
        len(unique_apps),
        len(sessions_any),
        len(sessions_threat),
        len(sessions_app),
        len(sessions_threat.union(sessions_app)),
        ""  # Comentarios
    ]

    for sev in SEV_PA:
        datos = threat_by_sev[sev]
        if datos["alerts"] > 0:
            res.extend([
                datos["alerts"],
                datos["alerts"],
                sum(len(s) for s in datos["session_ids"].values()),
                ", ".join(sorted(list(datos["ids"]))),
                len(datos["ids"]),
                len(datos["sessions"])
            ])
        else:
            res.extend([0, 0, 0, 0, 0, 0])

    for risk in RISK_PA:
        datos = app_by_risk[risk]
        if datos["alerts"] > 0:
            res.extend([
                datos["alerts"],
                datos["alerts"],
                sum(len(s) for s in datos["session_ids"].values()),
                ", ".join(sorted(list(datos["ids"]))),
                len(datos["ids"]),
                len(datos["sessions"])
            ])
        else:
            res.extend([0, 0, 0, 0, 0, 0])

    return res


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.getcwd()
    os.chdir(base_dir)

    print("=== PROCESADOR DE TRÁFICO LEGÍTIMO: PALO ALTO ===")

    ruta_legitimo = os.path.join("PA", "Legitimo")
    if not os.path.exists(ruta_legitimo):
        print(f"[!] ERROR: No existe la ruta {ruta_legitimo}.")
        return

    plantilla = "plantillaPA_legitimo.xlsx"
    nombre_salida = "Resultados_Legitimo_PA.xlsx"

    if not os.path.exists(plantilla):
        print(f"[!] ERROR: No se encuentra '{plantilla}'.")
        return

    print(f"[*] Escaneando logs en {ruta_legitimo} y generando {nombre_salida}...")

    valores_fila = procesar_logs_legitimo_pa(ruta_legitimo)

    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active

    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    fill_ips_color = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill_app_color = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')

    row_num = 3
    for col_num, val in enumerate(valores_fila, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.border = border_thin
        cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')

        if col_num >= 10 and col_num <= 39:
            cell.fill = fill_ips_color
        elif col_num >= 40:
            cell.fill = fill_app_color

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 35
    for c in range(1, 70):
        ws.column_dimensions[get_column_letter(c)].width = 16

    wb.save(nombre_salida)
    print(f"✅ ¡Hecho! Archivo guardado ordenadamente en: {nombre_salida}")


if __name__ == '__main__':
    main()