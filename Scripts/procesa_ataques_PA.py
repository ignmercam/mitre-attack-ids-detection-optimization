import pandas as pd
import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIGURACIÓN DE NIVELES (SEGÚN PLANTILLA)
# ==========================================
SEV_PA = ['1', '2', '3', '4', '5']
RISK_PA = ['1', '2', '3', '4', '5']


def analizar_fichero_log_pa(ruta_archivo):
    """Parsea un archivo .log de Palo Alto y extrae métricas detalladas por flujos."""
    sessions_any = set()
    sessions_threat = set()
    sessions_app = set()

    total_alerts = 0
    unique_threat_ids = set()
    unique_apps = set()

    # Estructuras internas por nivel de severidad/riesgo
    threat_by_sev = {s: {"alerts": 0, "ids": set(), "sessions": set()} for s in SEV_PA}
    app_by_risk = {r: {"alerts": 0, "ids": set(), "sessions": set()} for r in RISK_PA}

    with open(ruta_archivo, 'r', errors='ignore') as f:
        for line in f:
            # Extraer identificador de flujo (sessionid)
            match_session = re.search(r'sessionid="(\d+)"', line)
            session_id = match_session.group(1) if match_session else None

            # Extraer campos clave de PA mediante regex
            match_threat = re.search(r'threat_id="([^"]+)"', line)
            match_app = re.search(r'app="([^"]+)"', line)
            match_sev_num = re.search(r'severity_number="(\d+)"', line)
            match_risk_num = re.search(r'risk_of_app="(\d+)"', line)

            if match_threat or match_app:
                total_alerts += 1
                if session_id:
                    sessions_any.add(session_id)

                # Procesar motor IPS / Amenazas
                if match_threat:
                    t_id_raw = match_threat.group(1)

                    # CORRECCIÓN CRÍTICA: Extraer solo el número identificador dentro de los paréntesis
                    id_match = re.search(r'\((\d+)\)', t_id_raw)
                    if id_match:
                        t_id = id_match.group(1)
                    else:
                        # Por si acaso viene el número suelto sin paréntesis
                        id_match_alt = re.search(r'(\d+)', t_id_raw)
                        t_id = id_match_alt.group(1) if id_match_alt else t_id_raw

                    unique_threat_ids.add(t_id)
                    if session_id:
                        sessions_threat.add(session_id)
                    if match_sev_num:
                        sev = match_sev_num.group(1)
                        if sev in threat_by_sev:
                            threat_by_sev[sev]["alerts"] += 1
                            threat_by_sev[sev]["ids"].add(t_id)
                            if session_id:
                                threat_by_sev[sev]["sessions"].add(session_id)

                # Procesar motor Control de Aplicaciones
                if match_app:
                    app_name = match_app.group(1)
                    unique_apps.add(app_name)
                    if session_id:
                        sessions_app.add(session_id)
                    if match_risk_num:
                        risk = match_risk_num.group(1)
                        if risk in app_by_risk:
                            app_by_risk[risk]["alerts"] += 1
                            app_by_risk[risk]["ids"].add(app_name)
                            if session_id:
                                app_by_risk[risk]["sessions"].add(session_id)

    # ----------------------------------------------------
    # CONSTRUCCIÓN DE LA FILA SECUENCIAL (59 COLUMNAS TOTAL)
    # ----------------------------------------------------
    row_values = []
    row_values.append(os.path.basename(ruta_archivo))

    # Bloque General (8 columnas)
    row_values.append(total_alerts)
    row_values.append(len(unique_threat_ids))
    row_values.append(len(unique_apps))
    row_values.append(len(sessions_any))
    row_values.append(len(sessions_threat))
    row_values.append(len(sessions_app))
    row_values.append(len(sessions_any))
    row_values.append("")  # Comentarios

    # Bloques IPS severity (25 columnas)
    for sev in SEV_PA:
        row_values.append(threat_by_sev[sev]["alerts"])
        row_values.append(threat_by_sev[sev]["alerts"])
        row_values.append(", ".join(sorted(list(threat_by_sev[sev]["ids"]))))
        row_values.append(len(threat_by_sev[sev]["ids"]))
        row_values.append(len(threat_by_sev[sev]["sessions"]))

    # Bloques APP risk_of_app (25 columnas)
    for risk in RISK_PA:
        row_values.append(app_by_risk[risk]["alerts"])
        row_values.append(app_by_risk[risk]["alerts"])
        row_values.append(", ".join(sorted(list(app_by_risk[risk]["ids"]))))
        row_values.append(len(app_by_risk[risk]["ids"]))
        row_values.append(len(app_by_risk[risk]["sessions"]))

    return row_values


def generar_excel_plantilla_pa(directorio_logs, ruta_fichero_orden, nombre_salida):
    """Genera el Excel estructurado en bloques y ordenado según la lista oficial de ataques."""
    print(f"Cargando el orden oficial desde: {ruta_fichero_orden}...")
    if ruta_fichero_orden.endswith('.csv'):
        df_orden = pd.read_csv(ruta_fichero_orden)
    else:
        df_orden = pd.read_excel(ruta_fichero_orden)

    ordered_bases = [df_orden.columns[0]] + df_orden.iloc[:, 0].dropna().tolist()

    archivos = [os.path.join(directorio_logs, f) for f in os.listdir(directorio_logs) if f.endswith('.log')]
    if not archivos:
        print(f"No se encontraron archivos .log en la carpeta '{directorio_logs}'")
        return

    def obtener_clave_orden(ruta_completa):
        filename = os.path.basename(ruta_completa)
        for index, base in enumerate(ordered_bases):
            if base in filename:
                return (index, filename)
        return (len(ordered_bases), filename)

    archivos_ordenados = sorted(archivos, key=obtener_clave_orden)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = True
    ws.title = "Resultados PA"

    top_headers = [
        ("Archivo", 1),
        ('PA [IPS severity_number="Any"] || [App risk_of_app="Any"]', 8),
        ('PA [IPS severity_number="1"]', 5),
        ('PA [IPS severity_number="2"]', 5),
        ('PA [IPS severity_number="3"]', 5),
        ('PA [IPS severity_number="4"]', 5),
        ('PA [IPS severity_number="5"]', 5),
        ('PA [APP risk_of_app="1"]', 5),
        ('PA [APP risk_of_app="2"]', 5),
        ('PA [APP risk_of_app="3"]', 5),
        ('PA [APP risk_of_app="4"]', 5),
        ('PA [APP risk_of_app="5"]', 5)
    ]

    sub_any = ["#Alertas", "#threat_id", "#AppID", "#Flujos PA",
               "#Flujos con Ataques detectados (threatid)", "#Flujos con Ataques detectados (AppID)",
               "#Flujos con Ataques detectados (con threatid y/o appid)", "Comentarios detecciones"]
    sub_ips = ["#Alertas", "#Alertas/threat_id", "threat_ids (sin repetición)", "#threat_id (sin repetición)",
               "#Flujos con Ataques detectados (threatid)"]
    sub_app = ["#Alertas", "#Alertas/appid", "appids (sin repetición)", "#appid (sin repetición)",
               "#Flujos con Ataques detectados (appid)"]

    col_actual = 1
    for texto, span in top_headers:
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col_actual, end_row=1, end_column=col_actual + span - 1)
            ws.cell(row=1, column=col_actual, value=texto)
            col_actual += span
        else:
            ws.cell(row=1, column=col_actual, value=texto)
            col_actual += 1

    sub_cabeceras_totales = ["Archivo"] + sub_any
    for _ in range(5): sub_cabeceras_totales.extend(sub_ips)
    for _ in range(5): sub_cabeceras_totales.extend(sub_app)
    for idx, texto_sub in enumerate(sub_cabeceras_totales, start=1):
        ws.cell(row=2, column=idx, value=texto_sub)

    fill_top = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    fill_sub = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    font_white = Font(color='FFFFFF', bold=True, size=10)
    font_sub = Font(color='FFFFFF', bold=True, size=9)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for c in range(1, len(sub_cabeceras_totales) + 1):
        ws.cell(row=1, column=c).fill = fill_top
        ws.cell(row=1, column=c).font = font_white
        ws.cell(row=1, column=c).alignment = alignment_center
        ws.cell(row=2, column=c).fill = fill_sub
        ws.cell(row=2, column=c).font = font_sub
        ws.cell(row=2, column=c).alignment = alignment_center

    print("Procesando y ordenando los logs de Palo Alto con IDs numéricos...")
    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    fill_ips_color = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill_app_color = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')

    row_num = 3
    for archivo in archivos_ordenados:
        valores_fila = analizar_fichero_log_pa(archivo)

        for col_num, val in enumerate(valores_fila, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = border_thin
            cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')

            if col_num >= 10 and col_num <= 34:
                cell.fill = fill_ips_color
            elif col_num >= 35:
                cell.fill = fill_app_color

        row_num += 1

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 35
    for c in range(1, len(sub_cabeceras_totales) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    wb.save(nombre_salida)
    print(f"¡Listo! Archivo guardado con IDs numéricos limpios en: {nombre_salida}")


if __name__ == "__main__":
    CARPETA_LOGS = "./PA"
    FICHERO_ORDEN = "./orden.xlsx"
    EXCEL_SALIDA = "Analisis_Estructurado_PA_Final_Numerico.xlsx"

    generar_excel_plantilla_pa(CARPETA_LOGS, FICHERO_ORDEN, EXCEL_SALIDA)